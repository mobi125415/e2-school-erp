import shutil as file_shutil

import hashlib, hmac, secrets, base64
from pathlib import Path as FilePath
from flask import jsonify

import json, threading, time
from urllib import request as urllib_request
from urllib.error import HTTPError, URLError
from werkzeug.utils import secure_filename
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, send_from_directory, abort
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date
from pathlib import Path
from functools import wraps
import sys, threading, webbrowser
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.utils import ImageReader
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfgen import canvas
from reportlab.graphics.barcode import code128
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io, os, csv, json
import qrcode

def resource_path(relative):
    """Read bundled templates/static both in normal Python and PyInstaller EXE."""
    base=Path(getattr(sys,"_MEIPASS",Path(__file__).resolve().parent))
    return base/relative

def data_root():
    """Use writable storage locally and /tmp on Vercel/serverless."""
    if os.environ.get("VERCEL") or os.environ.get("AWS_LAMBDA_FUNCTION_NAME"):
        base=Path(os.environ.get("TMPDIR") or os.environ.get("TEMP") or "/tmp")
    elif sys.platform.startswith("win"):
        base=Path(os.environ.get("LOCALAPPDATA",Path.home()))
    else:
        base=Path.home()/".local"/"share"
    p=base/"E2Solutions"/"SchoolERP"
    p.mkdir(parents=True,exist_ok=True)
    return p

def recover_legacy_database(target):
    """Recover an older project-local database without overwriting current data."""
    if target.exists():
        return False
    candidates=[]
    try:
        here=Path(__file__).resolve().parent
        candidates += [
            here/"school_erp.db", here/"instance"/"school_erp.db",
            Path.cwd()/"school_erp.db", Path.cwd()/"instance"/"school_erp.db",
            here.parent/"school_erp.db", here.parent/"instance"/"school_erp.db"
        ]
    except Exception:
        pass
    # Only accept a non-empty SQLite file. Never overwrite an existing database.
    for candidate in candidates:
        try:
            if candidate.resolve() == target.resolve():
                continue
            if candidate.exists() and candidate.is_file() and candidate.stat().st_size > 8192:
                target.parent.mkdir(parents=True, exist_ok=True)
                file_shutil.copy2(candidate, target)
                print(f"[E2] Recovered legacy database: {candidate}")
                return True
        except Exception as exc:
            print("[E2] Legacy database recovery skipped:", exc)
    return False

DATA_DIR=data_root()
UPLOAD_DIR=DATA_DIR/"uploads"
UPLOAD_DIR.mkdir(parents=True,exist_ok=True)
DB_PATH=DATA_DIR/"school_erp.db"

app=Flask(__name__,template_folder=str(resource_path("templates")),static_folder=str(resource_path("static")))
app.config["MAX_CONTENT_LENGTH"]=500 * 1024 * 1024  # 500 MB learning-resource upload limit
app.config["SECRET_KEY"] = (
    os.environ.get("SECRET_KEY")
    or os.environ.get("E2_SECRET_KEY")
    or hashlib.sha256(
        ("E2-School-ERP:" + os.environ.get("VERCEL_URL","local")).encode("utf-8")
    ).hexdigest()
)
DATABASE_URL=os.environ.get("DATABASE_URL","").strip()
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL="postgresql://" + DATABASE_URL[len("postgres://"):]
app.config["SQLALCHEMY_DATABASE_URI"]=DATABASE_URL or f"sqlite:///{DB_PATH.as_posix()}"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"]=False
app.config["UPLOAD_FOLDER"]=str(UPLOAD_DIR)
db=SQLAlchemy(app)

@app.route("/manifest.webmanifest")
def pwa_manifest():
    return jsonify({
        "name": "E-2 School ERP",
        "short_name": "E-2 ERP",
        "start_url": "/dashboard",
        "display": "standalone",
        "background_color": "#f4f7fb",
        "theme_color": "#2458d3",
        "description": "E-2 School ERP mobile-ready school management system",
        "icons": [
            {"src": "/static/icons/icon-192.png", "sizes": "192x192", "type": "image/png", "purpose": "any maskable"},
            {"src": "/static/icons/icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any maskable"}
        ]
    })

@app.route("/user-logo/<path:filename>")
def user_logo(filename):
    from flask import send_from_directory
    return send_from_directory(UPLOAD_DIR,filename)

def e2_pdf_footer_style(styles):
    from reportlab.lib.styles import ParagraphStyle
    return ParagraphStyle("E2Footer", parent=styles["Normal"], fontSize=7, leading=8, textColor=colors.HexColor("#666666"), alignment=1, spaceBefore=5)

class Institute(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    name=db.Column(db.String(150),nullable=False)
    tagline=db.Column(db.String(180),default="")
    email=db.Column(db.String(150),unique=True,nullable=False)
    logo=db.Column(db.String(255))
    primary_color=db.Column(db.String(20),default="#2457d6")
    phone=db.Column(db.String(50),default="")
    address=db.Column(db.String(255),default="")
    active=db.Column(db.Boolean,default=True)

class User(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    institute_id=db.Column(db.Integer,db.ForeignKey("institute.id"),nullable=True)
    email=db.Column(db.String(150),unique=True,nullable=False)
    password_hash=db.Column(db.String(255),nullable=False)
    role=db.Column(db.String(30),default="school_admin")
    institute=db.relationship("Institute")

class Student(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    institute_id=db.Column(db.Integer,db.ForeignKey("institute.id"),nullable=False)
    admission_no=db.Column(db.String(50),nullable=False)
    name=db.Column(db.String(120),nullable=False)
    father_name=db.Column(db.String(120),default="")
    phone=db.Column(db.String(50),default="")
    class_name=db.Column(db.String(50),default="")
    section=db.Column(db.String(20),default="")
    monthly_fee=db.Column(db.Float,default=0)
    books_copies_amount=db.Column(db.Float,default=0)
    photo=db.Column(db.String(255),default="")
    status=db.Column(db.String(20),default="Active")
    joined_on=db.Column(db.Date,default=date.today)
    __table_args__=(db.UniqueConstraint("institute_id","admission_no",name="uq_inst_admission"),)

class StudentPortalAccount(db.Model):
    id=db.Column(db.Integer,primary_key=True); institute_id=db.Column(db.Integer,db.ForeignKey("institute.id"),nullable=False); student_id=db.Column(db.Integer,db.ForeignKey("student.id"),unique=True,nullable=False); password_hash=db.Column(db.String(255),nullable=False); active=db.Column(db.Boolean,default=True); student=db.relationship("Student")
class Lecture(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    institute_id=db.Column(db.Integer,db.ForeignKey("institute.id"),nullable=False)
    class_name=db.Column(db.String(80),nullable=False)
    subject=db.Column(db.String(120),nullable=False)
    chapter=db.Column(db.String(160),default="")
    title=db.Column(db.String(200),nullable=False)
    description=db.Column(db.Text,default="")
    content_type=db.Column(db.String(30),default="Video")
    video_url=db.Column(db.String(700),default="")
    file_name=db.Column(db.String(255),default="")
    created_on=db.Column(db.Date,default=date.today)

class Fee(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    institute_id=db.Column(db.Integer,db.ForeignKey("institute.id"),nullable=False)
    student_id=db.Column(db.Integer,db.ForeignKey("student.id"),nullable=False)
    month=db.Column(db.String(30),nullable=False)
    amount=db.Column(db.Float,default=0); discount=db.Column(db.Float,default=0); paid=db.Column(db.Float,default=0)
    paid_on=db.Column(db.Date); note=db.Column(db.String(255),default="")
    student=db.relationship("Student",backref="fees")

class Transaction(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    institute_id=db.Column(db.Integer,db.ForeignKey("institute.id"),nullable=False)
    kind=db.Column(db.String(20),nullable=False); category=db.Column(db.String(80),default="")
    description=db.Column(db.String(255),default=""); amount=db.Column(db.Float,default=0); txn_date=db.Column(db.Date,default=date.today)

class StaffSalary(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    institute_id=db.Column(db.Integer,db.ForeignKey("institute.id"),nullable=False)
    staff_name=db.Column(db.String(120),nullable=False)
    designation=db.Column(db.String(100),default="")
    month=db.Column(db.String(30),nullable=False)
    salary=db.Column(db.Float,default=0)
    paid=db.Column(db.Float,default=0)
    paid_on=db.Column(db.Date)
    note=db.Column(db.String(255),default="")


class TeacherProfile(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    institute_id=db.Column(db.Integer,db.ForeignKey("institute.id"),nullable=False)
    user_id=db.Column(db.Integer,db.ForeignKey("user.id"),unique=True)
    employee_no=db.Column(db.String(50),nullable=False)
    full_name=db.Column(db.String(140),nullable=False)
    father_husband_name=db.Column(db.String(140),default="")
    cnic=db.Column(db.String(30),default="")
    gender=db.Column(db.String(20),default="")
    dob=db.Column(db.Date)
    phone=db.Column(db.String(40),default="")
    whatsapp=db.Column(db.String(40),default="")
    email=db.Column(db.String(140),default="")
    address=db.Column(db.Text,default="")
    photo=db.Column(db.String(255),default="")
    joining_date=db.Column(db.Date)
    designation=db.Column(db.String(100),default="Teacher")
    qualification=db.Column(db.String(180),default="")
    university=db.Column(db.String(180),default="")
    specialization=db.Column(db.String(180),default="")
    experience=db.Column(db.String(100),default="")
    monthly_salary=db.Column(db.Float,default=0)
    emergency_contact=db.Column(db.String(80),default="")
    status=db.Column(db.String(20),default="Active")
    remarks=db.Column(db.Text,default="")
    user=db.relationship("User")

class TeacherAssignment(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    institute_id=db.Column(db.Integer,db.ForeignKey("institute.id"),nullable=False)
    teacher_user_id=db.Column(db.Integer,db.ForeignKey("user.id"),nullable=False)
    class_name=db.Column(db.String(50),nullable=False)
    subject=db.Column(db.String(80),default="")
    teacher=db.relationship("User")

class Attendance(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    institute_id=db.Column(db.Integer,db.ForeignKey("institute.id"),nullable=False)
    student_id=db.Column(db.Integer,db.ForeignKey("student.id"),nullable=False)
    attendance_date=db.Column(db.Date,default=date.today,nullable=False)
    status=db.Column(db.String(20),default="Present")
    marked_by=db.Column(db.Integer,db.ForeignKey("user.id"))
    student=db.relationship("Student")
    __table_args__=(db.UniqueConstraint("student_id","attendance_date",name="uq_student_attendance_day"),)

class Homework(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    institute_id=db.Column(db.Integer,db.ForeignKey("institute.id"),nullable=False)
    teacher_user_id=db.Column(db.Integer,db.ForeignKey("user.id"),nullable=False)
    class_name=db.Column(db.String(50),nullable=False)
    subject=db.Column(db.String(80),default="")
    title=db.Column(db.String(180),nullable=False)
    details=db.Column(db.Text,default="")
    due_date=db.Column(db.Date)
    created_on=db.Column(db.Date,default=date.today)

class ClassSubject(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    institute_id=db.Column(db.Integer,db.ForeignKey("institute.id"),nullable=False)
    class_name=db.Column(db.String(50),nullable=False)
    subject=db.Column(db.String(80),nullable=False)
    total_marks=db.Column(db.Float,default=100)

class ExamSession(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    institute_id=db.Column(db.Integer,db.ForeignKey("institute.id"),nullable=False)
    name=db.Column(db.String(120),nullable=False)
    class_name=db.Column(db.String(50),nullable=False)
    exam_date=db.Column(db.Date,default=date.today)

class ExamMark(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    institute_id=db.Column(db.Integer,db.ForeignKey("institute.id"),nullable=False)
    exam_session_id=db.Column(db.Integer,db.ForeignKey("exam_session.id"),nullable=False)
    student_id=db.Column(db.Integer,db.ForeignKey("student.id"),nullable=False)
    class_subject_id=db.Column(db.Integer,db.ForeignKey("class_subject.id"),nullable=False)
    obtained_marks=db.Column(db.Float,default=0)
    student=db.relationship("Student")
    class_subject=db.relationship("ClassSubject")

class Exam(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    institute_id=db.Column(db.Integer,db.ForeignKey("institute.id"),nullable=False)
    name=db.Column(db.String(120),nullable=False); class_name=db.Column(db.String(50),default="")
    subject=db.Column(db.String(80),default=""); total_marks=db.Column(db.Float,default=100); exam_date=db.Column(db.Date,default=date.today)

class Result(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    institute_id=db.Column(db.Integer,db.ForeignKey("institute.id"),nullable=False)
    exam_id=db.Column(db.Integer,db.ForeignKey("exam.id"),nullable=False)
    student_id=db.Column(db.Integer,db.ForeignKey("student.id"),nullable=False)
    obtained_marks=db.Column(db.Float,default=0)
    exam=db.relationship("Exam",backref="results"); student=db.relationship("Student",backref="results")

def _table_columns(table_name):
    """Portable SQLAlchemy schema inspection for SQLite and PostgreSQL."""
    from sqlalchemy import inspect
    try:
        return {c["name"] for c in inspect(db.engine).get_columns(table_name)}
    except Exception as exc:
        print(f"[ERP] Schema inspection failed for {table_name}: {exc!r}")
        return set()

def migrate_db():
    """Upgrade legacy Student columns without SQLite-only PRAGMA SQL."""
    try:
        cols=_table_columns("student")
        if not cols:
            return
        additions=[]
        if "books_copies_amount" not in cols:
            additions.append(("books_copies_amount","FLOAT DEFAULT 0"))
        if "photo" not in cols:
            additions.append(("photo","VARCHAR(255) DEFAULT ''"))
        for name,kind in additions:
            db.session.execute(db.text(f'ALTER TABLE student ADD COLUMN "{name}" {kind}'))
        if additions:
            db.session.commit()
    except Exception as exc:
        db.session.rollback()
        print("[ERP] Student migration skipped/failed:",repr(exc))

def migrate_institute_branding():
    try:
        cols=_table_columns("institute")
        if not cols:
            return
        if "tagline" not in cols:
            db.session.execute(db.text('ALTER TABLE institute ADD COLUMN "tagline" VARCHAR(180) DEFAULT \'\'' ))
            db.session.commit()
    except Exception as exc:
        db.session.rollback()
        print("[ERP] Institute branding migration skipped/failed:",repr(exc))

def current_inst():
    iid=session.get("institute_id")
    return Institute.query.get(iid) if iid else None

@app.context_processor
def inject():
    return {"inst":current_inst(),"year":datetime.now().year,"e2_contact":"+923010012627","is_master":session.get("role")=="superadmin"}

def login_required(fn):
    @wraps(fn)
    def w(*a,**kw):
        if not session.get("uid"): return redirect(url_for("login"))
        return fn(*a,**kw)
    return w

def school_required(fn):
    @wraps(fn)
    def w(*a,**kw):
        if not session.get("uid"): return redirect(url_for("login"))
        if session.get("role")=="superadmin": return redirect(url_for("master"))
        return fn(*a,**kw)
    return w

def teacher_required(fn):
    @wraps(fn)
    def w(*a,**kw):
        if not session.get("uid"): return redirect(url_for("login"))
        if session.get("role") not in ("teacher","school_admin"):
            return redirect(url_for("dashboard"))
        return fn(*a,**kw)
    return w

def student_portal_required(fn):
    @wraps(fn)
    def w(*a,**kw):
        if not session.get("student_id"): return redirect(url_for("login"))
        return fn(*a,**kw)
    return w

@app.route("/")
def startup_intro():
    return redirect(url_for("login"))

@app.route("/login",methods=["GET","POST"])
def login():
    if request.method=="POST":
        typ=request.form.get("login_type","admin")
        if typ=="student":
            inst=Institute.query.filter_by(email=request.form.get("institute_email","").strip().lower(),active=True).first()
            stu=Student.query.filter_by(institute_id=inst.id,admission_no=request.form.get("admission_no","").strip(),status="Active").first() if inst else None
            acc=StudentPortalAccount.query.filter_by(student_id=stu.id,active=True).first() if stu else None
            if acc and check_password_hash(acc.password_hash,request.form.get("password","")):
                session.clear(); session["student_id"]=stu.id; session["student_institute_id"]=inst.id; return redirect(url_for("student_portal"))
            flash("Invalid student login details.","danger")
        else:
            u=User.query.filter_by(email=request.form.get("email","").strip().lower()).first()
            if u and u.role!="teacher" and check_password_hash(u.password_hash,request.form.get("password","")):
                session.clear(); session["uid"]=u.id; session["role"]=u.role; session["institute_id"]=u.institute_id; return redirect(url_for("master") if u.role=="superadmin" else url_for("dashboard"))
            flash("Invalid administrator login details.","danger")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/master",methods=["GET","POST"])
@login_required
def master():
    if session.get("role")!="superadmin": return redirect(url_for("dashboard"))
    if request.method=="POST":
        try:
            name=request.form.get("name","").strip()
            email=request.form.get("email","").strip().lower()
            password=request.form.get("password","")
            if not name or not email or not password:
                flash("Institute name, login email and password are required.","danger")
                return redirect(url_for("master"))
            if Institute.query.filter_by(email=email).first() or User.query.filter_by(email=email).first():
                flash("This email is already in use. Please choose another login email.","danger")
                return redirect(url_for("master"))
            i=Institute(name=name,email=email,primary_color=request.form.get("primary_color","#2457d6"))
            db.session.add(i)
            db.session.flush()
            db.session.add(User(
                institute_id=i.id,
                email=email,
                password_hash=generate_password_hash(password),
                role="school_admin"
            ))
            db.session.commit()
            flash("New school created successfully.","success")
        except Exception as exc:
            db.session.rollback()
            print("[E2] New school creation failed:",repr(exc))
            flash("Could not create the school. Check the Vercel function log for the database error.","danger")
        return redirect(url_for("master"))
    return render_template("master.html",institutes=Institute.query.order_by(Institute.id.desc()).all())

@app.post("/master/institute/<int:iid>/toggle")
@login_required
def toggle_institute(iid):
    if session.get("role")!="superadmin": return redirect(url_for("dashboard"))
    i=Institute.query.get_or_404(iid); i.active=not i.active; db.session.commit(); return redirect(url_for("master"))

@app.post("/master/institute/<int:iid>/password")
@login_required
def reset_school_password(iid):
    if session.get("role")!="superadmin": return redirect(url_for("dashboard"))
    u=User.query.filter_by(institute_id=iid,role="school_admin").first_or_404()
    u.password_hash=generate_password_hash(request.form["password"]); db.session.commit()
    flash("School password changed.","success"); return redirect(url_for("master"))

@app.route("/dashboard")
@school_required
def dashboard():
    iid=session["institute_id"]
    cls=request.args.get("class_name","").strip()
    month=request.args.get("month","").strip()

    total_active=Student.query.filter_by(institute_id=iid,status="Active").count()

    students_q=Student.query.filter_by(institute_id=iid,status="Active")
    if cls:
        students_q=students_q.filter_by(class_name=cls)
    student_rows=students_q.all()
    sids=[s.id for s in student_rows]

    fq=Fee.query.filter(Fee.institute_id==iid)
    if sids:
        fq=fq.filter(Fee.student_id.in_(sids))
    elif cls:
        fq=fq.filter(db.text("1=0"))
    if month:
        fq=fq.filter(Fee.month==month)
    fee_rows=fq.all()

    latest={}
    for f in fee_rows:
        if f.student_id not in latest or f.id>latest[f.student_id].id:
            latest[f.student_id]=f

    paid_count=0
    unpaid_count=0
    due_amount=0
    for s in student_rows:
        f=latest.get(s.id)
        if f:
            due=max(0,(f.amount-f.discount)-f.paid)
            due_amount += due
            if due<=0 and f.paid>0:
                paid_count+=1
            else:
                unpaid_count+=1
        else:
            unpaid_count+=1
            due_amount += (s.monthly_fee+s.books_copies_amount)

    allfees=Fee.query.filter_by(institute_id=iid).all()
    tx=Transaction.query.filter_by(institute_id=iid).all()
    collected=sum(float(x.paid or 0) for x in allfees)

    # Lightweight live metrics for the professional dashboard. These are read-only
    # additions and do not change any existing ERP workflow or stored data.
    teacher_count=TeacherProfile.query.filter_by(institute_id=iid,status="Active").count()
    exam_count=ExamSession.query.filter_by(institute_id=iid).count()
    today_attendance=Attendance.query.filter_by(institute_id=iid,attendance_date=date.today()).all()
    present_today=sum(1 for a in today_attendance if (a.status or "").strip().lower()=="present")
    attendance_today_pct=round((present_today/len(today_attendance))*100,1) if today_attendance else 0
    profit=sum(x.amount for x in tx if x.kind=="Income")-sum(x.amount for x in tx if x.kind=="Expense")

    raw_classes=[r[0] for r in db.session.query(Student.class_name).filter_by(institute_id=iid,status="Active").all() if r[0]]
    seen=set()
    classes=[]
    for c in raw_classes:
        clean=" ".join(c.split())
        key=clean.casefold()
        if key not in seen:
            seen.add(key)
            classes.append(clean)
    classes.sort(key=lambda x:x.casefold())

    months=[r[0] for r in db.session.query(Fee.month).filter_by(institute_id=iid).distinct().order_by(Fee.id.desc()).all() if r[0]]
    recent_students=Student.query.filter_by(institute_id=iid).order_by(Student.id.desc()).limit(5).all()
    class_counts={}
    for s in Student.query.filter_by(institute_id=iid,status="Active").all():
        key=(s.class_name or "Unassigned").strip() or "Unassigned"
        class_counts[key]=class_counts.get(key,0)+1
    top_classes=sorted(class_counts.items(),key=lambda x:(-x[1],x[0]))[:6]
    fee_rate=round((collected/(collected+due))*100,1) if (collected+due)>0 else 0

    return render_template("dashboard.html",
        total_active=total_active,active=len(student_rows),
        old=Student.query.filter(Student.institute_id==iid,Student.status!="Active").count(),
        collected=sum(x.paid for x in allfees),due=due_amount,paid_count=paid_count,unpaid_count=unpaid_count,
        income=sum(x.amount for x in tx if x.kind=="Income"),expense=sum(x.amount for x in tx if x.kind=="Expense"),
        teacher_count=teacher_count,exam_count=exam_count,present_today=present_today,
        attendance_today_pct=attendance_today_pct,profit=profit,today_attendance_count=len(today_attendance),
        classes=classes,months=months,selected_class=cls,selected_month=month,
        top_classes=top_classes,recent_students=recent_students,fee_rate=fee_rate)
@app.route("/setup",methods=["GET","POST"])
@school_required
def setup():
    i=current_inst()
    if request.method=="POST":
        i.name=request.form["name"].strip(); i.primary_color=request.form.get("primary_color") or "#2457d6"
        i.tagline=request.form.get("tagline","").strip(); i.phone=request.form.get("phone",""); i.address=request.form.get("address","")
        f=request.files.get("logo")
        if f and f.filename:
            ext=Path(f.filename).suffix.lower()
            if ext in [".png",".jpg",".jpeg",".webp"]:
                name=f"logo_{i.id}{ext}"; f.save(Path(app.config["UPLOAD_FOLDER"])/name); i.logo=name
        db.session.commit(); flash("Branding updated.","success"); return redirect(url_for("setup"))
    return render_template("setup.html")

@app.route("/students",methods=["GET","POST"])
@school_required
def students():
    iid=session["institute_id"]
    if request.method=="POST":
        try:
            s=Student(institute_id=iid,admission_no=request.form["admission_no"].strip(),name=request.form["name"].strip(),
              father_name=request.form.get("father_name",""),phone=request.form.get("phone",""),class_name=request.form.get("class_name",""),
              section=request.form.get("section",""),monthly_fee=float(request.form.get("monthly_fee") or 0),books_copies_amount=float(request.form.get("books_copies_amount") or 0),status=request.form.get("status","Active"))
            db.session.add(s); db.session.flush()
            photo=request.files.get("photo")
            if photo and photo.filename:
                ext=Path(photo.filename).suffix.lower()
                if ext in [".png",".jpg",".jpeg",".webp"]:
                    filename=f"student_{iid}_{s.id}{ext}"
                    photo.save(UPLOAD_DIR/filename); s.photo=filename
            db.session.commit(); flash("Student added.","success")
        except Exception:
            db.session.rollback(); flash("Admission number must be unique in this school.","danger")
        return redirect(url_for("students"))
    cls=request.args.get("class_name","").strip(); q=request.args.get("q","").strip()
    query=Student.query.filter_by(institute_id=iid)
    if cls: query=query.filter_by(class_name=cls)
    if q: query=query.filter(db.or_(Student.name.ilike(f"%{q}%"),Student.admission_no.ilike(f"%{q}%"),Student.father_name.ilike(f"%{q}%"),Student.phone.ilike(f"%{q}%")))
    classes=[r[0] for r in db.session.query(Student.class_name).filter_by(institute_id=iid).distinct().order_by(Student.class_name).all() if r[0]]
    return render_template("students.html",students=query.order_by(Student.class_name,Student.name).all(),classes=classes,selected_class=cls,q=q)


@app.route("/left-students")
@school_required
def left_students():
    iid=session["institute_id"]
    cls=request.args.get("class_name","").strip()
    q=request.args.get("q","").strip()
    query=Student.query.filter(Student.institute_id==iid,Student.status.in_(["Left","Old"]))
    if cls: query=query.filter_by(class_name=cls)
    if q:
        query=query.filter(db.or_(Student.name.ilike(f"%{q}%"),Student.admission_no.ilike(f"%{q}%"),
                                 Student.father_name.ilike(f"%{q}%"),Student.phone.ilike(f"%{q}%")))
    raw=[r[0] for r in db.session.query(Student.class_name).filter(
        Student.institute_id==iid,Student.status.in_(["Left","Old"])).all() if r[0]]
    seen=set(); classes=[]
    for c in raw:
        clean=" ".join(c.split()); key=clean.casefold()
        if key not in seen: seen.add(key); classes.append(clean)
    classes.sort(key=lambda x:x.casefold())
    rows=query.order_by(Student.class_name,Student.name).all()
    return render_template("left_students.html",students=rows,classes=classes,selected_class=cls,q=q)

@app.route("/left-students.pdf")
@school_required
def left_students_pdf():
    iid=session["institute_id"]; i=current_inst()
    cls=request.args.get("class_name","").strip()
    q=request.args.get("q","").strip()
    query=Student.query.filter(Student.institute_id==iid,Student.status.in_(["Left","Old"]))
    if cls: query=query.filter_by(class_name=cls)
    if q:
        query=query.filter(db.or_(Student.name.ilike(f"%{q}%"),Student.admission_no.ilike(f"%{q}%"),
                                 Student.father_name.ilike(f"%{q}%"),Student.phone.ilike(f"%{q}%")))
    rows=query.order_by(Student.class_name,Student.name).all()
    out=io.BytesIO()
    doc=SimpleDocTemplate(out,pagesize=A4,rightMargin=18,leftMargin=18,topMargin=24,bottomMargin=24)
    st=getSampleStyleSheet(); story=[]
    if i.logo:
        from reportlab.platypus import Image
        logo_path=UPLOAD_DIR/i.logo
        if logo_path.exists():
            im=Image(str(logo_path),width=48,height=48); im.hAlign="CENTER"; story.append(im)
    story += [Paragraph(i.name,st["Title"]),
              Paragraph("Left / Old Students Record",st["Heading2"]),
              Paragraph(f"Class: {cls or 'All Classes'} | Total Students: {len(rows)}",st["Normal"]),
              Spacer(1,8)]
    data=[["Adm #","Student","Guardian","Phone","Class","Sec","Fee","Books","Status"]]
    for s in rows:
        data.append([s.admission_no,s.name,s.father_name,s.phone,s.class_name,s.section,
                     f"{s.monthly_fee:.0f}",f"{s.books_copies_amount:.0f}",s.status])
    table=Table(data,repeatRows=1,colWidths=[42,78,78,68,48,30,48,48,45])
    table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor(i.primary_color)),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),7.2),
        ("GRID",(0,0),(-1,-1),.35,colors.grey),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f5f7fb")])
    ]))
    story += [table,Spacer(1,10),
              Paragraph(f"<b>Total Left/Old Students: {len(rows)}</b>",st["Normal"]),
              Paragraph("Powered by E-2 Solutions | Contact: +923010012627",e2_pdf_footer_style(st))]
    doc.build(story); out.seek(0)
    return send_file(out,as_attachment=True,download_name=f"left_students_{cls or 'all'}.pdf",mimetype="application/pdf")

@app.post("/student/<int:sid>/status")
@school_required
def student_status(sid):
    s=Student.query.filter_by(id=sid,institute_id=session["institute_id"]).first_or_404()
    s.status=request.form["status"]; db.session.commit(); return redirect(url_for("students"))

def class_export_rows(class_name, fee_status="all"):
    iid=session["institute_id"]
    students=Student.query.filter_by(institute_id=iid,class_name=class_name).order_by(Student.name).all()
    rows=[]
    for s in students:
        f=Fee.query.filter_by(institute_id=iid,student_id=s.id).order_by(Fee.id.desc()).first()
        due=(max(0,(f.amount-f.discount)-f.paid) if f else (s.monthly_fee+s.books_copies_amount))
        status="Paid" if f and due<=0 and f.paid>0 else "Unpaid"
        if fee_status.lower() in ("paid","unpaid") and status.lower()!=fee_status.lower(): continue
        rows.append((s,status,due))
    return rows

@app.route("/class/<path:class_name>/xlsx")
@school_required
def class_xlsx(class_name):
    fee_status=request.args.get("fee_status","all")
    rows=class_export_rows(class_name,fee_status)
    wb=Workbook(); ws=wb.active; ws.title="Students"
    headers=["Admission No","Student Name","Father/Guardian","Phone","Class","Section","Monthly Fee","Books/Copies","Fee Status","Balance","Student Status"]
    ws.append(headers)
    for c in ws[1]:
        c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="2457D6"); c.alignment=Alignment(horizontal="center")
    for s,fs,due in rows: ws.append([s.admission_no,s.name,s.father_name,s.phone,s.class_name,s.section,s.monthly_fee,s.books_copies_amount,fs,due,s.status])
    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=min(max(len(str(x.value or "")) for x in col)+2,30)
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,as_attachment=True,download_name=f"{class_name}_{fee_status}_students.xlsx",mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/class/<path:class_name>/pdf")
@school_required
def class_pdf(class_name):
    iid=session["institute_id"]; i=current_inst(); fee_status=request.args.get("fee_status","all")
    rows=class_export_rows(class_name,fee_status)
    out=io.BytesIO(); doc=SimpleDocTemplate(out,pagesize=A4,rightMargin=18,leftMargin=18,topMargin=24,bottomMargin=24)
    st=getSampleStyleSheet(); story=[]
    if i.logo:
        from reportlab.platypus import Image
        logo_path=UPLOAD_DIR/i.logo
        if logo_path.exists():
            im=Image(str(logo_path),width=48,height=48); im.hAlign="CENTER"; story.append(im)
    story += [Paragraph(i.name,st["Title"]),Paragraph(f"Class: {class_name} | Fee Filter: {fee_status.title()}",st["Heading2"]),Spacer(1,8)]
    data=[["Adm #","Student","Guardian","Phone","Sec","Fee","Books","Fee Status","Balance"]]+[
        [s.admission_no,s.name,s.father_name,s.phone,s.section,f"{(s.monthly_fee+s.books_copies_amount):.0f}",f"{s.books_copies_amount:.0f}",fs,f"{due:.0f}"] for s,fs,due in rows]
    t=Table(data,repeatRows=1,colWidths=[40,75,75,66,28,38,38,50,42])
    t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor(i.primary_color)),("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),7.2),("GRID",(0,0),(-1,-1),.35,colors.grey),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f5f7fb")])]))
    story += [t,Spacer(1,12),Paragraph("Powered by E-2 Solutions | Contact: +923010012627",e2_pdf_footer_style(st))]
    doc.build(story); out.seek(0)
    return send_file(out,as_attachment=True,download_name=f"{class_name}_{fee_status}_students.pdf",mimetype="application/pdf")

@app.route("/student/<int:sid>/edit",methods=["GET","POST"])
@school_required
def edit_student(sid):
    s=Student.query.filter_by(id=sid,institute_id=session["institute_id"]).first_or_404()
    if request.method=="POST":
        try:
            s.admission_no=request.form["admission_no"].strip()
            s.name=request.form["name"].strip()
            s.father_name=request.form.get("father_name","")
            s.phone=request.form.get("phone","")
            s.class_name=request.form.get("class_name","")
            s.section=request.form.get("section","")
            s.monthly_fee=float(request.form.get("monthly_fee") or 0)
            s.books_copies_amount=float(request.form.get("books_copies_amount") or 0)
            s.status=request.form.get("status","Active")
            photo=request.files.get("photo")
            if photo and photo.filename:
                ext=Path(photo.filename).suffix.lower()
                if ext in [".png",".jpg",".jpeg",".webp"]:
                    filename=f"student_{s.institute_id}_{s.id}{ext}"
                    photo.save(UPLOAD_DIR/filename); s.photo=filename
            db.session.commit(); flash("Student updated.","success"); return redirect(url_for("students"))
        except Exception:
            db.session.rollback(); flash("Could not update student. Check admission number/details.","danger")
    return render_template("edit_student.html",s=s)


@app.route("/id-cards")
@school_required
def id_card_generator():
    iid=session["institute_id"]
    cls=request.args.get("class_name","").strip()
    classes=sorted({" ".join(x[0].split()) for x in db.session.query(Student.class_name).filter_by(institute_id=iid,status="Active").all() if x[0]},key=str.casefold)
    students=Student.query.filter_by(institute_id=iid,status="Active")
    if cls: students=students.filter_by(class_name=cls)
    return render_template("id_cards.html",classes=classes,students=students.order_by(Student.name).all(),selected_class=cls)

@app.route("/classes")
@school_required
def classes_page():
    iid=session["institute_id"]
    rows=[]
    for name,count in db.session.query(Student.class_name,db.func.count(Student.id)).filter_by(institute_id=iid,status="Active").group_by(Student.class_name).order_by(Student.class_name).all():
        if name: rows.append((name,count))
    return render_template("classes.html",rows=rows)

@app.route("/subjects")
@school_required
def subjects_page():
    iid=session["institute_id"]
    rows=ClassSubject.query.filter_by(institute_id=iid).order_by(ClassSubject.class_name,ClassSubject.subject).all()
    return render_template("subjects.html",rows=rows)

@app.route("/results")
@school_required
def results_page():
    iid=session["institute_id"]
    rows=ExamSession.query.filter_by(institute_id=iid).order_by(ExamSession.exam_date.desc()).all()
    return render_template("results.html",exam_sessions=rows)

@app.route("/expenses")
@school_required
def expenses_page():
    iid=session["institute_id"]
    rows=Transaction.query.filter_by(institute_id=iid,kind="Expense").order_by(Transaction.txn_date.desc(),Transaction.id.desc()).all()
    return render_template("expenses.html",rows=rows,total=sum(x.amount for x in rows))



@app.route("/student-portal-settings",methods=["GET","POST"])
@school_required
def student_portal_settings():
    iid=session["institute_id"]
    if request.method=="POST":
        sid=int(request.form["student_id"]); acc=StudentPortalAccount.query.filter_by(institute_id=iid,student_id=sid).first()
        if not acc: acc=StudentPortalAccount(institute_id=iid,student_id=sid,password_hash=generate_password_hash(request.form["password"]),active=True); db.session.add(acc)
        else:
            if request.form.get("password"): acc.password_hash=generate_password_hash(request.form["password"])
            acc.active=request.form.get("active")=="1"
        db.session.commit(); flash("Student portal login saved.","success"); return redirect(url_for("student_portal_settings"))
    students=Student.query.filter_by(institute_id=iid,status="Active").order_by(Student.class_name,Student.name).all(); accounts={a.student_id:a for a in StudentPortalAccount.query.filter_by(institute_id=iid).all()}; return render_template("student_portal_settings.html",students=students,accounts=accounts)

def save_upload(file_storage):
    """Save an administrator-uploaded learning resource safely and return stored filename."""
    if not file_storage or not getattr(file_storage, "filename", ""):
        return ""
    original=secure_filename(file_storage.filename)
    if not original:
        raise ValueError("Invalid file name.")
    ext=Path(original).suffix.lower()
    allowed={".mp4",".webm",".mov",".m4v",".pdf",".ppt",".pptx",".doc",".docx",".jpg",".jpeg",".png"}
    if ext not in allowed:
        raise ValueError("Unsupported file type: "+ext)
    upload_dir=Path(UPLOAD_DIR)
    upload_dir.mkdir(parents=True,exist_ok=True)
    stem=Path(original).stem[:80] or "resource"
    import uuid
    stored=f"{stem}_{uuid.uuid4().hex[:10]}{ext}"
    file_storage.save(str(upload_dir/stored))
    return stored

@app.route("/lectures",methods=["GET","POST"])
@school_required
def lectures():
    migrate_erp22_content_bank()
    iid=session["institute_id"]
    if request.method=="POST":
        try:
            class_name=request.form.get("class_name","").strip()
            subject=request.form.get("subject","").strip()
            title=request.form.get("title","").strip()
            if not class_name or not subject or not title:
                flash("Class, subject and title are required.","danger")
                return redirect(url_for("lectures"))
            fn=""
            f=request.files.get("lecture_file")
            if f and f.filename:
                ext=Path(f.filename).suffix.lower()
                allowed={".mp4",".webm",".mov",".m4v",".pdf",".ppt",".pptx",".doc",".docx",".jpg",".jpeg",".png"}
                if ext not in allowed:
                    flash("Unsupported file type.","danger")
                    return redirect(url_for("lectures"))
                fn=save_upload(f)
            db.session.add(Lecture(
                institute_id=iid,class_name=class_name,subject=subject,
                chapter=request.form.get("chapter","").strip(),
                title=title,description=request.form.get("description","").strip(),
                content_type=request.form.get("content_type","Video"),
                video_url=request.form.get("video_url","").strip(),
                file_name=fn))
            db.session.commit()
            flash("Learning content published successfully.","success")
        except Exception as e:
            db.session.rollback()
            flash("Could not publish content: "+str(e),"danger")
        return redirect(url_for("lectures"))

    classes=[x[0] for x in db.session.query(Student.class_name).filter(
        Student.institute_id==iid).distinct().order_by(Student.class_name).all() if x[0]]
    rows=Lecture.query.filter_by(institute_id=iid).order_by(
        Lecture.class_name,Lecture.subject,Lecture.id.desc()).all()
    return render_template("lectures.html",classes=classes,rows=rows)

@app.post("/lectures/<int:lid>/delete")
@school_required
def lecture_delete(lid):
    x=Lecture.query.filter_by(id=lid,institute_id=session["institute_id"]).first_or_404()
    db.session.delete(x)
    db.session.commit()
    flash("Learning content deleted.","success")
    return redirect(url_for("lectures"))

@app.route("/lecture-media/<int:lid>")
def lecture_media(lid):
    x=Lecture.query.get_or_404(lid)
    admin_ok=session.get("authenticated") and session.get("institute_id")==x.institute_id
    student_ok=session.get("student_id") and session.get("student_institute_id")==x.institute_id
    if not (admin_ok or student_ok):
        abort(403)
    if not x.file_name:
        abort(404)
    return send_from_directory(UPLOAD_DIR,x.file_name,as_attachment=False)

@app.route("/attendance-admin",methods=["GET","POST"])
@school_required
def attendance_admin():
    iid=session["institute_id"]; classes=[x[0] for x in db.session.query(Student.class_name).filter_by(institute_id=iid,status="Active").distinct().all() if x[0]]; cls=request.values.get("class_name",""); ds=request.values.get("attendance_date") or date.today().isoformat(); ad=date.fromisoformat(ds); students=Student.query.filter_by(institute_id=iid,class_name=cls,status="Active").order_by(Student.name).all() if cls else []
    if request.method=="POST":
        for st in students:
            a=Attendance.query.filter_by(student_id=st.id,attendance_date=ad).first() or Attendance(institute_id=iid,student_id=st.id,attendance_date=ad,marked_by=session["uid"]); db.session.add(a); a.status=request.form.get(f"status_{st.id}","Present")
        db.session.commit(); return redirect(url_for("attendance_admin",class_name=cls,attendance_date=ds))
    existing={a.student_id:a.status for a in Attendance.query.filter_by(institute_id=iid,attendance_date=ad).all()}; return render_template("attendance_admin.html",classes=classes,students=students,selected_class=cls,attendance_date=ds,existing=existing)
def _student_context():
    migrate_erp22_content_bank()
    iid=session["student_institute_id"]; sid=session["student_id"]
    st=Student.query.filter_by(id=sid,institute_id=iid,status="Active").first_or_404()
    inst=Institute.query.get(iid)
    return iid,sid,st,inst

def _student_results(iid,sid,st):
    rows=[]
    try:
        exams=ExamSession.query.filter_by(institute_id=iid,class_name=st.class_name).order_by(ExamSession.exam_date.desc()).all()
        for e in exams:
            marks=ExamMark.query.filter_by(institute_id=iid,exam_session_id=e.id,student_id=sid).all()
            if marks:
                obt=sum(float(m.obtained_marks or 0) for m in marks)
                total=sum(float(m.class_subject.total_marks or 0) for m in marks if m.class_subject)
                rows.append((e,marks,obt,total))
    except Exception:
        pass
    return rows

@app.route("/student-portal")
@student_portal_required
def student_portal():
    iid,sid,st,inst=_student_context()
    attendance=Attendance.query.filter_by(institute_id=iid,student_id=sid).all()
    present=sum(1 for a in attendance if str(a.status).lower()=="present")
    total=len(attendance); pct=round(present*100/total,1) if total else 0
    lectures=Lecture.query.filter_by(institute_id=iid,class_name=st.class_name).all()
    results=_student_results(iid,sid,st)
    return render_template("student/dashboard.html",student=st,institute=inst,attendance_pct=pct,
        present=present,total=total,lecture_count=len(lectures),result_count=len(results),active="dashboard")

@app.route("/student/lectures")
@student_portal_required
def student_lectures():
    iid,sid,st,inst=_student_context()
    lectures=Lecture.query.filter_by(institute_id=iid,class_name=st.class_name).order_by(Lecture.subject,Lecture.chapter,Lecture.id.desc()).all()
    subjects={}
    for x in lectures:
        subject_name=(x.subject or "General").strip()
        subjects.setdefault(subject_name,[]).append(x)
    selected_subject=(request.args.get("subject") or "").strip()
    if selected_subject not in subjects:
        selected_subject=""
    return render_template("student/lectures.html",student=st,institute=inst,subjects=subjects,
        selected_subject=selected_subject,active="lectures")

@app.route("/student/results")
@student_portal_required
def student_results():
    iid,sid,st,inst=_student_context()
    return render_template("student/results.html",student=st,institute=inst,results=_student_results(iid,sid,st),active="results")

@app.route("/student/attendance")
@student_portal_required
def student_attendance():
    iid,sid,st,inst=_student_context()
    rows=Attendance.query.filter_by(institute_id=iid,student_id=sid).order_by(Attendance.attendance_date.desc()).all()
    present=sum(1 for a in rows if str(a.status).lower()=="present"); total=len(rows)
    pct=round(present*100/total,1) if total else 0
    return render_template("student/attendance.html",student=st,institute=inst,rows=rows,present=present,total=total,pct=pct,active="attendance")

@app.route("/student/profile")
@student_portal_required
def student_profile():
    iid,sid,st,inst=_student_context()
    return render_template("student/profile.html",student=st,institute=inst,active="profile")

@app.route("/student/logout")
def student_logout(): session.clear(); return redirect(url_for("login"))
@app.route("/teachers")
@school_required
def teachers():
    rows=TeacherProfile.query.filter_by(institute_id=session["institute_id"]).order_by(TeacherProfile.full_name).all()
    return render_template("teachers.html",rows=rows)

@app.route("/teachers/add",methods=["GET","POST"])
@school_required
def teacher_add():
    iid=session["institute_id"]
    if request.method=="POST":
        email=request.form["email"].strip().lower()
        if User.query.filter_by(email=email).first():
            flash("This login email already exists.","danger"); return redirect(url_for("teacher_add"))
        photo=""; f=request.files.get("photo")
        if f and f.filename: photo=save_upload(f)
        u=User(institute_id=iid,email=email,password_hash=generate_password_hash(request.form["password"]),role="teacher")
        db.session.add(u); db.session.flush()
        def dt(n):
            v=request.form.get(n,"").strip()
            return date.fromisoformat(v) if v else None
        t=TeacherProfile(institute_id=iid,user_id=u.id,employee_no=request.form["employee_no"].strip(),full_name=request.form["full_name"].strip(),
          father_husband_name=request.form.get("father_husband_name","").strip(),cnic=request.form.get("cnic","").strip(),gender=request.form.get("gender",""),
          dob=dt("dob"),phone=request.form.get("phone","").strip(),whatsapp=request.form.get("whatsapp","").strip(),email=email,address=request.form.get("address","").strip(),
          photo=photo,joining_date=dt("joining_date"),designation=request.form.get("designation","Teacher").strip(),qualification=request.form.get("qualification","").strip(),
          university=request.form.get("university","").strip(),specialization=request.form.get("specialization","").strip(),experience=request.form.get("experience","").strip(),
          monthly_salary=float(request.form.get("monthly_salary") or 0),emergency_contact=request.form.get("emergency_contact","").strip(),
          status=request.form.get("status","Active"),remarks=request.form.get("remarks","").strip())
        db.session.add(t); db.session.commit(); flash("Teacher profile and login created.","success"); return redirect(url_for("teachers"))
    return render_template("teacher_form.html",t=None)

@app.route("/teachers/<int:tid>/edit",methods=["GET","POST"])
@school_required
def teacher_edit(tid):
    t=TeacherProfile.query.filter_by(id=tid,institute_id=session["institute_id"]).first_or_404()
    if request.method=="POST":
        new_email=request.form["email"].strip().lower()
        if User.query.filter(User.email==new_email,User.id!=t.user_id).first():
            flash("Email already used by another user.","danger"); return redirect(url_for("teacher_edit",tid=tid))
        for fld in ["employee_no","full_name","father_husband_name","cnic","gender","phone","whatsapp","address","designation","qualification","university","specialization","experience","emergency_contact","status","remarks"]:
            setattr(t,fld,request.form.get(fld,"").strip())
        t.email=new_email; t.monthly_salary=float(request.form.get("monthly_salary") or 0)
        for fld in ("dob","joining_date"):
            v=request.form.get(fld,"").strip(); setattr(t,fld,date.fromisoformat(v) if v else None)
        f=request.files.get("photo")
        if f and f.filename: t.photo=save_upload(f)
        u=User.query.get(t.user_id)
        if u:
            u.email=new_email
            if request.form.get("password","").strip(): u.password_hash=generate_password_hash(request.form["password"])
        db.session.commit(); flash("Teacher updated.","success"); return redirect(url_for("teachers"))
    return render_template("teacher_form.html",t=t)

@app.route("/teachers/<int:tid>")
@school_required
def teacher_view(tid):
    t=TeacherProfile.query.filter_by(id=tid,institute_id=session["institute_id"]).first_or_404()
    assignments=TeacherAssignment.query.filter_by(institute_id=t.institute_id,teacher_user_id=t.user_id).all()
    return render_template("teacher_view.html",t=t,assignments=assignments)

@app.post("/teachers/<int:tid>/delete")
@school_required
def teacher_delete(tid):
    t=TeacherProfile.query.filter_by(id=tid,institute_id=session["institute_id"]).first_or_404(); uid=t.user_id
    TeacherAssignment.query.filter_by(institute_id=t.institute_id,teacher_user_id=uid).delete()
    Homework.query.filter_by(institute_id=t.institute_id,teacher_user_id=uid).delete()
    db.session.delete(t); db.session.flush(); u=User.query.get(uid)
    if u: db.session.delete(u)
    db.session.commit(); flash("Teacher and login deleted.","success"); return redirect(url_for("teachers"))

@app.route("/teachers/<int:tid>/profile.pdf")
@school_required
def teacher_profile_pdf(tid):
    t=TeacherProfile.query.filter_by(id=tid,institute_id=session["institute_id"]).first_or_404()
    i=current_inst(); out=io.BytesIO(); c=canvas.Canvas(out,pagesize=A4); W,H=A4; brand=colors.HexColor(i.primary_color or "#5B22C6")
    c.setFillColor(brand); c.rect(0,H-105,W,105,fill=1,stroke=0)
    if i.logo and (UPLOAD_DIR/i.logo).exists(): c.drawImage(str(UPLOAD_DIR/i.logo),38,H-88,60,60,preserveAspectRatio=True,mask="auto")
    c.setFillColor(colors.white); c.setFont("Helvetica-Bold",20); c.drawString(115,H-55,(i.name or "Institute")[:38])
    c.setFont("Helvetica",10); c.drawString(115,H-75,(i.tagline or "Teacher Profile")[:65])
    c.setFillColor(colors.HexColor("#222")); c.setFont("Helvetica-Bold",18); c.drawString(40,H-145,"TEACHER PROFILE")
    if t.photo and (UPLOAD_DIR/t.photo).exists(): c.drawImage(str(UPLOAD_DIR/t.photo),W-145,H-270,100,120,preserveAspectRatio=True,mask="auto")
    fields=[("Employee No.",t.employee_no),("Full Name",t.full_name),("Father / Husband",t.father_husband_name),("CNIC",t.cnic),("Gender",t.gender),("DOB",t.dob),
      ("Phone",t.phone),("WhatsApp",t.whatsapp),("Email",t.email),("Joining Date",t.joining_date),("Designation",t.designation),("Qualification",t.qualification),
      ("University / Board",t.university),("Specialization",t.specialization),("Experience",t.experience),("Monthly Salary",f"{t.monthly_salary:,.0f}"),
      ("Emergency Contact",t.emergency_contact),("Status",t.status)]
    y=H-185
    for lab,val in fields:
        c.setFont("Helvetica-Bold",9); c.drawString(45,y,lab); c.setFont("Helvetica",10); c.drawString(165,y,str(val or "-")[:55]); y-=27
    c.setFillColor(brand); c.rect(0,0,W,35,fill=1,stroke=0); c.setFillColor(colors.white); c.setFont("Helvetica-Bold",8)
    c.drawCentredString(W/2,13,"Powered by E-2 Solutions | +923010012627"); c.save(); out.seek(0)
    return send_file(out,as_attachment=True,download_name=f"{t.employee_no}_{t.full_name}_Profile.pdf",mimetype="application/pdf")

@app.route("/teacher")
@teacher_required
def teacher_portal():
    iid=session["institute_id"]; uid=session["uid"]
    assignments=TeacherAssignment.query.filter_by(institute_id=iid,teacher_user_id=uid).order_by(TeacherAssignment.class_name).all()
    classes=sorted({a.class_name for a in assignments})
    student_count=Student.query.filter(Student.institute_id==iid,Student.status=="Active",Student.class_name.in_(classes)).count() if classes else 0
    today=date.today()
    marked=Attendance.query.filter_by(institute_id=iid,marked_by=uid,attendance_date=today).count()
    homework=Homework.query.filter_by(institute_id=iid,teacher_user_id=uid).order_by(Homework.id.desc()).limit(8).all()
    return render_template("teacher_portal.html",assignments=assignments,student_count=student_count,marked=marked,homework=homework,today=today)

@app.route("/teacher/class/<path:class_name>")
@teacher_required
def teacher_class(class_name):
    iid=session["institute_id"]; uid=session["uid"]
    if session.get("role")=="teacher" and not TeacherAssignment.query.filter_by(institute_id=iid,teacher_user_id=uid,class_name=class_name).first():
        flash("This class is not assigned to your account.","danger"); return redirect(url_for("teacher_portal"))
    students=Student.query.filter_by(institute_id=iid,class_name=class_name,status="Active").order_by(Student.name).all()
    return render_template("teacher_class.html",students=students,class_name=class_name,today=date.today())

@app.post("/teacher/attendance/<path:class_name>")
@teacher_required
def teacher_attendance(class_name):
    iid=session["institute_id"]; uid=session["uid"]
    if session.get("role")=="teacher" and not TeacherAssignment.query.filter_by(institute_id=iid,teacher_user_id=uid,class_name=class_name).first_or_404():
        return redirect(url_for("teacher_portal"))
    ad=date.fromisoformat(request.form.get("attendance_date") or date.today().isoformat())
    students=Student.query.filter_by(institute_id=iid,class_name=class_name,status="Active").all()
    for st in students:
        status=request.form.get(f"status_{st.id}","Present")
        row=Attendance.query.filter_by(student_id=st.id,attendance_date=ad).first()
        if not row:
            row=Attendance(institute_id=iid,student_id=st.id,attendance_date=ad)
            db.session.add(row)
        row.status=status; row.marked_by=uid
    db.session.commit(); flash(f"Attendance saved for {class_name}.","success")
    return redirect(url_for("teacher_class",class_name=class_name))

@app.route("/teacher/homework",methods=["GET","POST"])
@teacher_required
def teacher_homework():
    iid=session["institute_id"]; uid=session["uid"]
    assignments=TeacherAssignment.query.filter_by(institute_id=iid,teacher_user_id=uid).order_by(TeacherAssignment.class_name).all()
    if request.method=="POST":
        cls=request.form["class_name"]
        if session.get("role")=="teacher" and not any(a.class_name==cls for a in assignments):
            flash("Class not assigned.","danger"); return redirect(url_for("teacher_homework"))
        due=request.form.get("due_date")
        db.session.add(Homework(institute_id=iid,teacher_user_id=uid,class_name=cls,subject=request.form.get("subject","").strip(),
            title=request.form["title"].strip(),details=request.form.get("details","").strip(),due_date=date.fromisoformat(due) if due else None))
        db.session.commit(); flash("Homework posted.","success"); return redirect(url_for("teacher_homework"))
    rows=Homework.query.filter_by(institute_id=iid,teacher_user_id=uid).order_by(Homework.id.desc()).all()
    return render_template("teacher_homework.html",assignments=assignments,rows=rows)

@app.route("/teacher-assignments",methods=["GET","POST"])
@school_required
def teacher_assignments():
    if session.get("role")!="school_admin": return redirect(url_for("teacher_portal"))
    iid=session["institute_id"]
    if request.method=="POST":
        db.session.add(TeacherAssignment(institute_id=iid,teacher_user_id=int(request.form["teacher_user_id"]),
            class_name=request.form["class_name"].strip(),subject=request.form.get("subject","").strip()))
        db.session.commit(); flash("Teacher assignment saved.","success"); return redirect(url_for("teacher_assignments"))
    teachers=User.query.filter_by(institute_id=iid,role="teacher").order_by(User.email).all()
    classes=[r[0] for r in db.session.query(Student.class_name).filter_by(institute_id=iid,status="Active").distinct().order_by(Student.class_name).all() if r[0]]
    rows=TeacherAssignment.query.filter_by(institute_id=iid).order_by(TeacherAssignment.class_name).all()
    return render_template("teacher_assignments.html",teachers=teachers,classes=classes,rows=rows)

@app.post("/teacher-assignments/<int:aid>/delete")
@school_required
def teacher_assignment_delete(aid):
    a=TeacherAssignment.query.filter_by(id=aid,institute_id=session["institute_id"]).first_or_404()
    db.session.delete(a); db.session.commit(); flash("Assignment removed.","success")
    return redirect(url_for("teacher_assignments"))

@app.route("/users",methods=["GET","POST"])
@school_required
def users_page():
    iid=session["institute_id"]
    if request.method=="POST":
        try:
            email=request.form["email"].strip().lower()
            password=request.form["password"]
            role=request.form.get("role","staff")
            if role not in ("school_admin","teacher","accountant","staff"): role="staff"
            db.session.add(User(institute_id=iid,email=email,password_hash=generate_password_hash(password),role=role))
            db.session.commit(); flash("User created successfully.","success")
        except Exception:
            db.session.rollback(); flash("Email already exists or details are invalid.","danger")
        return redirect(url_for("users_page"))
    return render_template("users.html",users=User.query.filter_by(institute_id=iid).order_by(User.id).all())

@app.post("/users/<int:uid>/password")
@school_required
def user_password(uid):
    u=User.query.filter_by(id=uid,institute_id=session["institute_id"]).first_or_404()
    u.password_hash=generate_password_hash(request.form["password"]); db.session.commit()
    flash("Password updated.","success"); return redirect(url_for("users_page"))

@app.post("/users/<int:uid>/delete")
@school_required
def user_delete(uid):
    if uid==session.get("uid"):
        flash("You cannot delete your own logged-in account.","danger"); return redirect(url_for("users_page"))
    u=User.query.filter_by(id=uid,institute_id=session["institute_id"]).first_or_404()
    db.session.delete(u); db.session.commit(); flash("User deleted.","success"); return redirect(url_for("users_page"))




@app.route("/cloud-status")
@school_required
def cloud_status_page():
    cloud_db=bool(os.environ.get("DATABASE_URL"))
    return render_template("cloud_status.html",cloud_db=cloud_db,
        production_secret=bool(os.environ.get("E2_SECRET_KEY")),
        port=os.environ.get("PORT","5000"))

@app.route("/health")
def cloud_health():
    try:
        db.session.execute(db.text("SELECT 1"))
        return {"status":"ok","service":"E-2 Solutions School ERP"},200
    except Exception as e:
        return {"status":"error","service":"E-2 Solutions School ERP"},503

@app.route("/mobile-test")
@school_required
def mobile_test_page():
    import socket
    try:
        sock=socket.socket(socket.AF_INET,socket.SOCK_DGRAM)
        sock.connect(("8.8.8.8",80))
        lan_ip=sock.getsockname()[0]
        sock.close()
    except Exception:
        lan_ip="YOUR-PC-IP"
    return render_template("mobile_test.html",lan_ip=lan_ip,port=int(os.environ.get("E2_PORT","5190")))

@app.route("/promotion",methods=["GET","POST"])
@school_required
def promotion():
    iid=session["institute_id"]
    classes=[x[0] for x in db.session.query(Student.class_name).filter_by(
        institute_id=iid,status="Active").distinct().order_by(Student.class_name).all() if x[0]]
    source=(request.values.get("source_class") or "").strip()
    target=(request.values.get("target_class") or "").strip()
    students=[]
    if source:
        students=Student.query.filter_by(institute_id=iid,class_name=source,status="Active").order_by(Student.name).all()
    if request.method=="POST" and request.form.get("action")=="promote":
        ids=[int(x) for x in request.form.getlist("student_ids") if x.isdigit()]
        if not source or not target:
            flash("Select source and next class.","danger")
        elif source==target:
            flash("Source and next class cannot be the same.","danger")
        elif not ids:
            flash("Select at least one student to promote.","danger")
        else:
            selected=Student.query.filter(Student.institute_id==iid,Student.id.in_(ids),
                Student.class_name==source,Student.status=="Active").all()
            for st in selected:
                st.class_name=target
            db.session.commit()
            flash(f"{len(selected)} student(s) promoted from {source} to {target}.","success")
            return redirect(url_for("promotion",source_class=target))
    return render_template("promotion.html",classes=classes,source=source,target=target,students=students)

@app.route("/database-security")
@school_required
def database_security():
    dbpath=current_database_path()
    db_size=dbpath.stat().st_size if dbpath and dbpath.exists() else 0
    cfg=WhatsAppConfig.query.filter_by(institute_id=session["institute_id"]).first()
    encrypted_token=bool(cfg and cfg.access_token and str(cfg.access_token).startswith("E2ENC1:"))
    return render_template("database_security.html",
        key_exists=MASTER_KEY_FILE.exists(), encrypted_token=encrypted_token,
        db_size=db_size, db_name=dbpath.name if dbpath else "Unknown")

@app.post("/database-security/protect-existing")
@school_required
def database_security_protect_existing():
    cfg=WhatsAppConfig.query.filter_by(institute_id=session["institute_id"]).first()
    changed=0
    if cfg and cfg.access_token and not str(cfg.access_token).startswith("E2ENC1:"):
        cfg.access_token=protect_secret(cfg.access_token); changed+=1
    db.session.commit()
    flash(f"Security update completed. {changed} existing secret(s) encrypted.","success")
    return redirect(url_for("database_security"))

@app.route("/backups")
@school_required
def backups_page():
    iid=session["institute_id"]
    backup_dir=DATA_DIR/"backups"; backup_dir.mkdir(parents=True,exist_ok=True)
    files=sorted(backup_dir.glob(f"institute_{iid}_*.json"),reverse=True)
    return render_template("backups.html",files=files[:30])

@app.post("/backups/create")
@school_required
def create_backup():
    iid=session["institute_id"]; i=current_inst()
    def rows(model):
        result=[]
        for obj in model.query.filter_by(institute_id=iid).all():
            d={}
            for col in obj.__table__.columns:
                v=getattr(obj,col.name)
                if isinstance(v,(date,datetime)): v=v.isoformat()
                d[col.name]=v
            result.append(d)
        return result
    payload={"version":"ERP14","created_at":datetime.now().isoformat(),
             "institute":{"id":i.id,"name":i.name,"email":i.email,"phone":i.phone,"address":i.address,"primary_color":i.primary_color},
             "students":rows(Student),"fees":rows(Fee),"transactions":rows(Transaction),"salaries":rows(StaffSalary),
             "subjects":rows(ClassSubject),"exam_sessions":rows(ExamSession),"exam_marks":rows(ExamMark)}
    backup_dir=DATA_DIR/"backups"; backup_dir.mkdir(parents=True,exist_ok=True)
    name=f"institute_{iid}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    (backup_dir/name).write_text(json.dumps(payload,ensure_ascii=False,indent=2),encoding="utf-8")
    flash("Backup created successfully.","success"); return redirect(url_for("backups_page"))

@app.route("/backups/download/<path:name>")
@school_required
def download_backup(name):
    safe=Path(name).name
    if not safe.startswith(f"institute_{session['institute_id']}_"): abort(403)
    return send_from_directory(DATA_DIR/"backups",safe,as_attachment=True)

@app.route("/certificate",methods=["GET","POST"])
@school_required
def certificate_page():
    iid=session["institute_id"]
    students=Student.query.filter_by(institute_id=iid).order_by(Student.name).all()
    return render_template("certificate.html",students=students)

@app.post("/certificate.pdf")
@school_required
def certificate_pdf():
    iid=session["institute_id"]; i=current_inst()
    s=Student.query.filter_by(id=int(request.form["student_id"]),institute_id=iid).first_or_404()
    title=request.form.get("title","Certificate of Achievement").strip()
    body=request.form.get("body","This certificate is proudly presented in recognition of excellent performance and conduct.").strip()
    out=io.BytesIO(); c=canvas.Canvas(out,pagesize=landscape(A4)); w,h=landscape(A4)
    brand=colors.HexColor(i.primary_color or "#5b20c8")
    c.setStrokeColor(brand); c.setLineWidth(5); c.rect(22,22,w-44,h-44)
    c.setLineWidth(1); c.rect(32,32,w-64,h-64)
    if i.logo and (UPLOAD_DIR/i.logo).exists(): c.drawImage(str(UPLOAD_DIR/i.logo),w/2-38,h-125,76,76,preserveAspectRatio=True,mask="auto")
    c.setFillColor(brand); c.setFont("Helvetica-Bold",27); c.drawCentredString(w/2,h-165,title[:55])
    c.setFillColor(colors.black); c.setFont("Helvetica",12); c.drawCentredString(w/2,h-195,"This is to certify that")
    c.setFont("Helvetica-Bold",25); c.drawCentredString(w/2,h-235,s.name[:45])
    c.setFont("Helvetica",12); c.drawCentredString(w/2,h-260,f"Admission No. {s.admission_no}  |  Class {s.class_name} {s.section}".strip())
    text=c.beginText(110,h-305); text.setFont("Helvetica",11); text.setLeading(17)
    for line in [body[x:x+95] for x in range(0,len(body),95)]: text.textLine(line)
    c.drawText(text)
    c.line(110,95,260,95); c.line(w-260,95,w-110,95); c.setFont("Helvetica",9)
    c.drawCentredString(185,80,"Class Teacher"); c.drawCentredString(w-185,80,"Principal")
    c.setFont("Helvetica",8); c.drawCentredString(w/2,52,f"{i.address} | {i.phone}")
    c.setFont("Helvetica-Bold",7); c.drawCentredString(w/2,38,"Powered by E-2 Solutions | +923010012627")
    c.save(); out.seek(0)
    return send_file(out,as_attachment=True,download_name=f"{s.name}_certificate.pdf",mimetype="application/pdf")

@app.route("/admit-card")
@school_required
def admit_card_page():
    iid=session["institute_id"]
    exams=ExamSession.query.filter_by(institute_id=iid).order_by(ExamSession.exam_date.desc()).all()
    students=Student.query.filter_by(institute_id=iid,status="Active").order_by(Student.name).all()
    return render_template("admit_card.html",exams=exams,students=students)

@app.get("/admit-card.pdf")
@school_required
def admit_card_pdf():
    iid=session["institute_id"]; i=current_inst()
    s=Student.query.filter_by(id=int(request.args["student_id"]),institute_id=iid).first_or_404()
    e=ExamSession.query.filter_by(id=int(request.args["exam_id"]),institute_id=iid).first_or_404()
    subjects=ClassSubject.query.filter_by(institute_id=iid,class_name=e.class_name).order_by(ClassSubject.id).all()
    out=io.BytesIO(); doc=SimpleDocTemplate(out,pagesize=A4,rightMargin=35,leftMargin=35,topMargin=30,bottomMargin=30)
    st=getSampleStyleSheet(); story=[]
    if i.logo and (UPLOAD_DIR/i.logo).exists():
        im=Image(str(UPLOAD_DIR/i.logo),width=55,height=55); im.hAlign="CENTER"; story.append(im)
    story += [Paragraph(i.name,st["Title"]),Paragraph("EXAMINATION ADMIT CARD",st["Heading2"]),Spacer(1,8)]
    info=[["Student",s.name,"Admission No.",s.admission_no],["Class",f"{s.class_name} {s.section}","Exam",e.name],["Exam Date",str(e.exam_date),"Contact",s.phone]]
    tb=Table(info,colWidths=[80,160,90,160]); tb.setStyle(TableStyle([("GRID",(0,0),(-1,-1),.4,colors.grey),("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("FONTNAME",(2,0),(2,-1),"Helvetica-Bold"),("BACKGROUND",(0,0),(0,-1),colors.HexColor("#f1f3f8")),("BACKGROUND",(2,0),(2,-1),colors.HexColor("#f1f3f8")),("PADDING",(0,0),(-1,-1),7)]))
    story += [tb,Spacer(1,14),Paragraph("Subjects",st["Heading3"])]
    sd=[["#","Subject","Total Marks"]]+[[n,x.subject,f"{x.total_marks:.0f}"] for n,x in enumerate(subjects,1)]
    ts=Table(sd,repeatRows=1,colWidths=[40,340,100]); ts.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor(i.primary_color)),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("GRID",(0,0),(-1,-1),.4,colors.grey),("PADDING",(0,0),(-1,-1),7)]))
    story += [ts,Spacer(1,35),Table([["____________________","____________________"],["Student Signature","Principal / Controller"]],colWidths=[250,250],style=[("ALIGN",(0,0),(-1,-1),"CENTER")]),Spacer(1,20),Paragraph(f"{i.address} | {i.phone}",st["Normal"]),Paragraph("Powered by E-2 Solutions | Contact: +923010012627",e2_pdf_footer_style(st))]
    doc.build(story); out.seek(0)
    return send_file(out,as_attachment=True,download_name=f"{s.name}_{e.name}_admit_card.pdf",mimetype="application/pdf")

@app.route("/student/<int:sid>/id-card")
@school_required
def student_id_card(sid):
    s=Student.query.filter_by(id=sid,institute_id=session["institute_id"]).first_or_404()
    return render_template("student_id_card.html",s=s)


def draw_professional_id_card(c,s,i,x,y,w,h):
    from urllib.parse import quote
    from reportlab.graphics.barcode import code128
    from reportlab.lib.utils import ImageReader
    sx,sy=w/500.0,h/278.0
    X=lambda v:x+v*sx; Y=lambda v:y+v*sy; W=lambda v:v*sx; H=lambda v:v*sy
    fs=lambda v:max(3.8,v*min(sx,sy))
    purple=colors.HexColor(i.primary_color or "#5B22C6")
    deep=colors.HexColor("#211044"); violet=colors.HexColor("#7138D2")
    # shell
    c.setFillColor(colors.white); c.setStrokeColor(colors.HexColor("#B9BDC8")); c.setLineWidth(max(.4,min(sx,sy)))
    c.roundRect(X(2),Y(2),W(496),H(274),W(15),fill=1,stroke=1)
    # header
    c.setFillColor(deep); c.roundRect(X(3),Y(208),W(494),H(67),W(15),fill=1,stroke=0); c.rect(X(3),Y(208),W(494),H(20),fill=1,stroke=0)
    c.setFillColor(violet); ph=c.beginPath(); ph.moveTo(X(300),Y(208)); ph.curveTo(X(380),Y(218),X(445),Y(248),X(497),Y(274)); ph.lineTo(X(497),Y(208)); ph.close(); c.drawPath(ph,fill=1,stroke=0)
    c.setFillColor(purple); ph=c.beginPath(); ph.moveTo(X(370),Y(208)); ph.curveTo(X(430),Y(217),X(470),Y(234),X(497),Y(251)); ph.lineTo(X(497),Y(208)); ph.close(); c.drawPath(ph,fill=1,stroke=0)
    if i.logo and (UPLOAD_DIR/i.logo).exists(): c.drawImage(str(UPLOAD_DIR/i.logo),X(24),Y(217),W(52),H(52),preserveAspectRatio=True,mask="auto")
    c.setFillColor(colors.white); c.setFont("Helvetica-Bold",fs(25)); c.drawString(X(92),Y(241),(i.name or "INSTITUTE").upper()[:28])
    c.setFont("Helvetica",fs(9)); c.drawString(X(94),Y(224),(i.tagline or "Quality Education, Bright Future")[:58])
    # photo
    c.setFillColor(colors.HexColor("#FAFAFC")); c.setStrokeColor(purple); c.setLineWidth(max(.5,1.5*min(sx,sy))); c.roundRect(X(18),Y(62),W(102),H(132),W(10),fill=1,stroke=1)
    if s.photo and (UPLOAD_DIR/s.photo).exists(): c.drawImage(str(UPLOAD_DIR/s.photo),X(22),Y(66),W(94),H(124),preserveAspectRatio=True,mask="auto")
    else:
        c.setFillColor(colors.grey); c.setFont("Helvetica-Bold",fs(8)); c.drawCentredString(X(69),Y(125),"STUDENT PHOTO")
    # details
    yy=183
    for lab,val,ic in [("Student Name",s.name,"S"),("Father / Guardian",s.father_name,"G"),("Class / Section",f"{s.class_name} {s.section}".strip(),"C"),("Admission No.",s.admission_no,"#"),("Contact",s.phone,"T")]:
        c.setFillColor(purple); c.circle(X(140),Y(yy),W(7),fill=1,stroke=0); c.setFillColor(colors.white); c.setFont("Helvetica-Bold",fs(6)); c.drawCentredString(X(140),Y(yy-2),ic)
        c.setFillColor(colors.HexColor("#30323A")); c.setFont("Helvetica",fs(8.5)); c.drawString(X(154),Y(yy-3),lab)
        c.setFillColor(colors.black); c.setFont("Helvetica-Bold",fs(10.5)); c.drawString(X(248),Y(yy-3),str(val or "")[:22])
        c.setStrokeColor(colors.HexColor("#DFE1E7")); c.line(X(154),Y(yy-10),X(342),Y(yy-10)); yy-=25
    c.setStrokeColor(colors.HexColor("#C9CBD2")); c.line(X(354),Y(66),X(354),Y(190))
    # barcode
    bc=code128.Code128(str(s.admission_no or s.id),barHeight=H(22),barWidth=max(.3,.72*sx)); bc.drawOn(c,X(154),Y(49))
    # whatsapp QR
    digits="".join(ch for ch in (i.phone or "") if ch.isdigit())
    if digits.startswith("0"): digits="92"+digits[1:]
    msg=f"{i.name}\n{i.tagline or ''}\nAddress: {i.address}\nStudent: {s.name}\nAdmission No: {s.admission_no}\nClass: {s.class_name} {s.section}"
    wa=f"https://wa.me/{digits}?text={quote(msg)}" if digits else f"https://wa.me/?text={quote(msg)}"
    qr=qrcode.QRCode(version=None,box_size=5,border=1); qr.add_data(wa); qr.make(fit=True)
    qi=qr.make_image(fill_color="black",back_color="white"); qb=io.BytesIO(); qi.save(qb,format="PNG"); qb.seek(0)
    c.setFillColor(colors.white); c.setStrokeColor(colors.HexColor("#E2E3E8")); c.roundRect(X(376),Y(111),W(96),H(92),W(8),fill=1,stroke=1); c.drawImage(ImageReader(qb),X(383),Y(118),W(82),H(80),mask="auto")
    c.setFillColor(colors.HexColor("#18A65B")); c.roundRect(X(381),Y(77),W(86),H(28),W(7),fill=1,stroke=0); c.setFillColor(colors.white); c.setFont("Helvetica-Bold",fs(10)); c.drawCentredString(X(424),Y(91),"WhatsApp"); c.setFont("Helvetica",fs(5.8)); c.drawCentredString(X(424),Y(82),"Scan to Chat")
    # signature
    c.setStrokeColor(colors.HexColor("#777A84")); c.line(X(205),Y(43),X(287),Y(43)); c.setFillColor(colors.HexColor("#333333")); c.setFont("Helvetica",fs(6)); c.drawCentredString(X(246),Y(34),"Principal")
    # layered wave footer
    c.setFillColor(purple); pw=c.beginPath(); pw.moveTo(X(3),Y(57)); pw.curveTo(X(95),Y(79),X(165),Y(50),X(255),Y(50)); pw.curveTo(X(350),Y(50),X(430),Y(66),X(497),Y(73)); pw.lineTo(X(497),Y(19)); pw.lineTo(X(3),Y(19)); pw.close(); c.drawPath(pw,fill=1,stroke=0)
    c.setFillColor(deep); pw=c.beginPath(); pw.moveTo(X(3),Y(47)); pw.curveTo(X(90),Y(67),X(170),Y(42),X(255),Y(42)); pw.curveTo(X(350),Y(42),X(430),Y(55),X(497),Y(62)); pw.lineTo(X(497),Y(19)); pw.lineTo(X(3),Y(19)); pw.close(); c.drawPath(pw,fill=1,stroke=0)
    c.setFillColor(colors.white); c.setFont("Helvetica",fs(6.8)); c.drawString(X(28),Y(34),(i.address or "Institute Address")[:62]); c.drawRightString(X(472),Y(34),"Issue Date  "+date.today().strftime("%d-%b-%Y"))
    c.setFillColor(purple); c.rect(X(3),Y(3),W(494),H(18),fill=1,stroke=0); c.setFillColor(colors.white); c.setFont("Helvetica-Bold",fs(8)); c.drawCentredString(X(250),Y(9),"Powered by E-2 Solutions  |  +923010012627")

@app.route("/student/<int:sid>/id-card.pdf")
@school_required
def student_id_card_pdf(sid):
    s=Student.query.filter_by(id=sid,institute_id=session["institute_id"]).first_or_404(); i=current_inst()
    out=io.BytesIO(); c=canvas.Canvas(out,pagesize=(500,278)); draw_professional_id_card(c,s,i,0,0,500,278); c.save(); out.seek(0)
    return send_file(out,as_attachment=True,download_name=f"{s.admission_no}_{s.name}_ID_Card.pdf",mimetype="application/pdf")

@app.route("/class/<path:class_name>/id-cards.pdf")
@school_required
def class_id_cards_pdf(class_name):
    iid=session["institute_id"]; i=current_inst()
    students=Student.query.filter_by(institute_id=iid,class_name=class_name,status="Active").order_by(Student.name).all()
    out=io.BytesIO(); c=canvas.Canvas(out,pagesize=A4); pw,ph=A4; cw,ch=252,140
    pos=[(32,ph-32-ch),(311,ph-32-ch),(32,ph-190-ch),(311,ph-190-ch),(32,ph-348-ch),(311,ph-348-ch),(32,ph-506-ch),(311,ph-506-ch)]
    for n,s in enumerate(students):
        if n and n%8==0: c.showPage()
        x,y=pos[n%8]; draw_professional_id_card(c,s,i,x,y,cw,ch)
    c.save(); out.seek(0)
    return send_file(out,as_attachment=True,download_name=f"{class_name}_Professional_ID_Cards.pdf",mimetype="application/pdf")

@app.post("/student/<int:sid>/delete")
@school_required
def delete_student(sid):
    s=Student.query.filter_by(id=sid,institute_id=session["institute_id"]).first_or_404()
    # Delete dependent records first.
    Fee.query.filter_by(institute_id=session["institute_id"],student_id=s.id).delete()
    Result.query.filter_by(institute_id=session["institute_id"],student_id=s.id).delete()
    db.session.delete(s); db.session.commit(); flash("Student deleted.","success")
    return redirect(url_for("students"))

@app.post("/fee/<int:fid>/delete")
@school_required
def delete_fee(fid):
    f=Fee.query.filter_by(id=fid,institute_id=session["institute_id"]).first_or_404()
    db.session.delete(f); db.session.commit(); flash("Fee entry deleted.","success")
    return redirect(url_for("fees"))

@app.route("/fee/<int:fid>/edit",methods=["GET","POST"])
@school_required
def edit_fee(fid):
    f=Fee.query.filter_by(id=fid,institute_id=session["institute_id"]).first_or_404()
    if request.method=="POST":
        f.month=request.form["month"]
        f.amount=float(request.form.get("amount") or 0)
        f.discount=float(request.form.get("discount") or 0)
        f.paid=float(request.form.get("paid") or 0)
        f.paid_on=date.today() if f.paid>0 else None
        f.note=request.form.get("note","")
        db.session.commit(); flash("Fee entry updated.","success"); return redirect(url_for("fees"))
    return render_template("edit_fee.html",f=f)

@app.get("/api/students")
@school_required
def api_students():
    iid=session["institute_id"]
    cls=request.args.get("class_name","").strip()
    q=request.args.get("q","").strip()
    # Fee balance is month-specific. Payments from July must never reduce August, etc.
    fee_month=request.args.get("month","").strip() or datetime.now().strftime("%B %Y")
    sq=Student.query.filter_by(institute_id=iid,status="Active")
    if cls: sq=sq.filter_by(class_name=cls)
    if q: sq=sq.filter(Student.name.ilike(f"%{q}%"))
    data=[]
    for s in sq.order_by(Student.name).limit(20).all():
        month_rows=Fee.query.filter_by(institute_id=iid,student_id=s.id,month=fee_month).all()
        paid_this_month=sum(float(x.paid or 0) for x in month_rows)
        discount_this_month=sum(float(x.discount or 0) for x in month_rows)
        total=float(s.monthly_fee or 0)+float(s.books_copies_amount or 0)
        remaining=max(0,total-paid_this_month-discount_this_month)
        data.append({"id":s.id,"name":s.name,"admission_no":s.admission_no,"class_name":s.class_name,
                     "monthly_fee":float(s.monthly_fee or 0),"books":float(s.books_copies_amount or 0),
                     "total":total,"paid_this_month":paid_this_month,
                     "discount_this_month":discount_this_month,"remaining":remaining,"month":fee_month})
    return {"students":data}

@app.route("/fees",methods=["GET","POST"])
@school_required
def fees():
    iid=session["institute_id"]
    current_month=datetime.now().strftime("%B %Y")
    if request.method=="POST":
        s=Student.query.filter_by(id=int(request.form["student_id"]),institute_id=iid).first_or_404()
        paid=float(request.form.get("paid") or 0)
        total_decided=s.monthly_fee+s.books_copies_amount
        amount=float(request.form.get("amount") or total_decided)
        f=Fee(institute_id=iid,student_id=s.id,month=request.form.get("month") or current_month,amount=amount,
              discount=float(request.form.get("discount") or 0),paid=paid,paid_on=date.today() if paid>0 else None,
              note=request.form.get("note","Monthly Fee" if paid>=amount else "Partial Fee"))
        db.session.add(f); db.session.commit(); flash("Fee saved.","success"); return redirect(url_for("fees"))
    cls=request.args.get("class_name","").strip(); q=request.args.get("q","").strip(); fee_status=request.args.get("fee_status","all")
    fq=Fee.query.filter_by(institute_id=iid)
    if cls: fq=fq.join(Student).filter(Student.class_name==cls)
    if q: fq=fq.join(Student, Fee.student_id==Student.id).filter(Student.name.ilike(f"%{q}%")) if not cls else fq.filter(Student.name.ilike(f"%{q}%"))
    fees_all=fq.order_by(Fee.id.desc()).all()
    if fee_status=="paid": fees_all=[f for f in fees_all if max(0,f.amount-f.discount-f.paid)<=0 and f.paid>0]
    elif fee_status=="unpaid": fees_all=[f for f in fees_all if max(0,f.amount-f.discount-f.paid)>0]
    classes=[r[0] for r in db.session.query(Student.class_name).filter_by(institute_id=iid,status="Active").distinct().order_by(Student.class_name).all() if r[0]]
    return render_template("fees.html",fees=fees_all,classes=classes,selected_class=cls,q=q,fee_status=fee_status,current_month=current_month)
@app.route("/fee-slip/<int:fid>")
@school_required
def fee_slip(fid):
    return render_template("fee_slip.html",fee=Fee.query.filter_by(id=fid,institute_id=session["institute_id"]).first_or_404())




SECURITY_DIR=FilePath(app.root_path)/"instance"
SECURITY_DIR.mkdir(parents=True,exist_ok=True)
MASTER_KEY_FILE=SECURITY_DIR/"e2_master.key"

def _master_key():
    if MASTER_KEY_FILE.exists():
        return MASTER_KEY_FILE.read_bytes()
    key=secrets.token_bytes(32)
    MASTER_KEY_FILE.write_bytes(key)
    try:
        import os
        os.chmod(MASTER_KEY_FILE,0o600)
    except Exception: pass
    return key

def _keystream(key,nonce,length):
    out=b""; counter=0
    while len(out)<length:
        out += hmac.new(key,nonce+counter.to_bytes(8,"big"),hashlib.sha256).digest()
        counter+=1
    return out[:length]

def protect_secret(value):
    if not value: return ""
    if str(value).startswith("E2ENC1:"): return str(value)
    key=_master_key(); nonce=secrets.token_bytes(16); raw=str(value).encode("utf-8")
    stream=_keystream(key,nonce,len(raw))
    cipher=bytes(a^b for a,b in zip(raw,stream))
    tag=hmac.new(key,nonce+cipher,hashlib.sha256).digest()
    return "E2ENC1:"+base64.urlsafe_b64encode(nonce+tag+cipher).decode("ascii")

def reveal_secret(value):
    if not value: return ""
    value=str(value)
    if not value.startswith("E2ENC1:"): return value
    try:
        blob=base64.urlsafe_b64decode(value[7:].encode("ascii"))
        nonce,tag,cipher=blob[:16],blob[16:48],blob[48:]
        key=_master_key()
        good=hmac.new(key,nonce+cipher,hashlib.sha256).digest()
        if not hmac.compare_digest(tag,good): return ""
        stream=_keystream(key,nonce,len(cipher))
        return bytes(a^b for a,b in zip(cipher,stream)).decode("utf-8")
    except Exception:
        return ""

class WhatsAppConfig(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    institute_id=db.Column(db.Integer,db.ForeignKey("institute.id"),unique=True,nullable=False)
    phone_number_id=db.Column(db.String(120),default="")
    access_token=db.Column(db.Text,default="")
    api_version=db.Column(db.String(30),default="v23.0")
    template_name=db.Column(db.String(120),default="fee_reminder")
    template_language=db.Column(db.String(20),default="en")
    enabled=db.Column(db.Boolean,default=False)
    reminder_day=db.Column(db.Integer,default=5)
    last_run_month=db.Column(db.String(20),default="")

class FeeReminderLog(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    institute_id=db.Column(db.Integer,db.ForeignKey("institute.id"),nullable=False)
    fee_id=db.Column(db.Integer,db.ForeignKey("fee.id"),nullable=False)
    student_id=db.Column(db.Integer,db.ForeignKey("student.id"),nullable=False)
    reminder_month=db.Column(db.String(20),nullable=False)
    phone=db.Column(db.String(40),default="")
    status=db.Column(db.String(30),default="Pending")
    message_id=db.Column(db.String(255),default="")
    error=db.Column(db.Text,default="")
    sent_at=db.Column(db.DateTime,default=datetime.utcnow)


def institute_brand_phone(inst):
    if not inst: return ""
    if getattr(inst,"phone",None): return str(getattr(inst,"phone")).strip()
    return ""

def normalize_pk_phone(value):
    phone="".join(ch for ch in (value or "") if ch.isdigit())
    if phone.startswith("0") and len(phone)>=10: phone="92"+phone[1:]
    return phone

def send_whatsapp_fee_template(cfg, phone, student, fee, due, institute):
    # The visible school identity comes from the logged-in institute branding.
    # Actual WhatsApp "From" identity is controlled by Meta Phone Number ID and
    # must correspond to this institute's registered WhatsApp Business number.
    branding_phone=institute_brand_phone(institute)
    # Template body parameters must match the approved Meta template:
    # {{1}} student, {{2}} admission, {{3}} class, {{4}} month,
    # {{5}} due amount, {{6}} school.
    url=f"https://graph.facebook.com/{cfg.api_version}/{cfg.phone_number_id}/messages"
    payload={
      "messaging_product":"whatsapp","recipient_type":"individual","to":phone,"type":"template",
      "template":{"name":cfg.template_name,"language":{"code":cfg.template_language},
        "components":[{"type":"body","parameters":[
          {"type":"text","text":student.name or ""},
          {"type":"text","text":student.admission_no or ""},
          {"type":"text","text":student.class_name or ""},
          {"type":"text","text":fee.month or ""},
          {"type":"text","text":f"{due:,.0f}"},
          {"type":"text","text":institute.name or "School"},
          {"type":"text","text":branding_phone or ""}]}]}}
    data=json.dumps(payload).encode("utf-8")
    req=urllib_request.Request(url,data=data,method="POST",headers={
      "Authorization":f"Bearer {reveal_secret(cfg.access_token)}","Content-Type":"application/json"})
    try:
        with urllib_request.urlopen(req,timeout=25) as r:
            result=json.loads(r.read().decode("utf-8"))
            mid=((result.get("messages") or [{}])[0]).get("id","")
            return True,mid,""
    except HTTPError as e:
        try: err=e.read().decode("utf-8")
        except Exception: err=str(e)
        return False,"",err[:1500]
    except Exception as e:
        return False,"",str(e)[:1500]

def run_monthly_fee_reminders(force=False, institute_id=None):
    now=datetime.now()
    configs=WhatsAppConfig.query.filter_by(enabled=True)
    if institute_id: configs=configs.filter_by(institute_id=institute_id)
    sent=failed=skipped=0
    for cfg in configs.all():
        if not force and now.day!=int(cfg.reminder_day or 5): continue
        run_key=now.strftime("%Y-%m")
        if not force and cfg.last_run_month==run_key: continue
        inst=Institute.query.get(cfg.institute_id)
        fees=Fee.query.filter_by(institute_id=cfg.institute_id).all()
        for fee in fees:
            due=max(0,float(fee.amount or 0)-float(fee.discount or 0)-float(fee.paid or 0))
            if due<=0: continue
            st=fee.student
            if not st: continue
            phone=normalize_pk_phone(st.phone)
            if not phone:
                skipped+=1; continue
            # one reminder per fee record per calendar month
            old=FeeReminderLog.query.filter_by(institute_id=cfg.institute_id,fee_id=fee.id,
                reminder_month=run_key,status="Sent").first()
            if old and not force:
                skipped+=1; continue
            ok,mid,err=send_whatsapp_fee_template(cfg,phone,st,fee,due,inst)
            db.session.add(FeeReminderLog(institute_id=cfg.institute_id,fee_id=fee.id,
                student_id=st.id,reminder_month=run_key,phone=phone,
                status="Sent" if ok else "Failed",message_id=mid,error=err))
            db.session.commit()
            if ok: sent+=1
            else: failed+=1
        if not force:
            cfg.last_run_month=run_key
            db.session.commit()
    return sent,failed,skipped

_scheduler_started=False
def start_fee_reminder_scheduler():
    global _scheduler_started
    if _scheduler_started: return
    _scheduler_started=True
    def worker():
        while True:
            try:
                with app.app_context():
                    run_monthly_fee_reminders()
            except Exception as e:
                print("[AUTO FEE REMINDER]",e)
            time.sleep(3600)  # hourly check; sends only on configured day and once/month
    threading.Thread(target=worker,daemon=True,name="fee-reminder-scheduler").start()


BACKUP_DIR=FilePath(app.root_path)/"backups"
BACKUP_DIR.mkdir(parents=True,exist_ok=True)

def current_database_path():
    uri=app.config.get("SQLALCHEMY_DATABASE_URI","")
    if uri.startswith("sqlite:///"):
        raw=uri.replace("sqlite:///","",1)
        path=FilePath(raw)
        if not path.is_absolute(): path=FilePath(app.root_path)/path
        return path.resolve()
    return None

def make_database_backup(label="auto"):
    dbpath=current_database_path()
    if not dbpath or not dbpath.exists():
        return None,"Database file not found or database is not SQLite."
    stamp=datetime.now().strftime("%Y%m%d_%H%M%S")
    safe="".join(ch for ch in label if ch.isalnum() or ch in "-_")[:25] or "backup"
    dest=BACKUP_DIR/f"school_erp_{safe}_{stamp}.db"
    # SQLite online backup API produces a consistent backup while ERP is running.
    import sqlite3
    srccon=sqlite3.connect(str(dbpath))
    dstcon=sqlite3.connect(str(dest))
    try: srccon.backup(dstcon)
    finally:
        dstcon.close(); srccon.close()
    return dest,None

def cleanup_old_backups(keep=30):
    files=sorted(BACKUP_DIR.glob("school_erp_*.db"),key=lambda x:x.stat().st_mtime,reverse=True)
    for f in files[keep:]:
        try:f.unlink()
        except Exception:pass

def backup_rows():
    rows=[]
    for f in sorted(BACKUP_DIR.glob("school_erp_*.db"),key=lambda x:x.stat().st_mtime,reverse=True):
        st=f.stat()
        rows.append({"name":f.name,"size":st.st_size,"modified":datetime.fromtimestamp(st.st_mtime)})
    return rows

_backup_scheduler_started=False
def start_backup_scheduler():
    global _backup_scheduler_started
    if _backup_scheduler_started:return
    _backup_scheduler_started=True
    def worker():
        last_day=""
        while True:
            try:
                now=datetime.now()
                key=now.strftime("%Y-%m-%d")
                # Daily automatic backup after 18:00 local time, once per day.
                if now.hour>=18 and key!=last_day:
                    with app.app_context():
                        path,err=make_database_backup("daily")
                        if path:
                            cleanup_old_backups(30)
                            print("[AUTO BACKUP]",path.name)
                        elif err: print("[AUTO BACKUP ERROR]",err)
                    last_day=key
            except Exception as e: print("[AUTO BACKUP ERROR]",e)
            time.sleep(1800)
    threading.Thread(target=worker,daemon=True,name="automatic-backup").start()

@app.route("/backups")
@school_required
def automatic_backups_page():
    return render_template("backups.html",rows=backup_rows(),backup_dir=str(BACKUP_DIR))

@app.post("/backups/create")
@school_required
def backup_create():
    path,err=make_database_backup("manual")
    if err: flash(err,"danger")
    else:
        cleanup_old_backups(30)
        flash("Backup created successfully: "+path.name,"success")
    return redirect(url_for("automatic_backups_page"))

@app.route("/backups/download/<path:name>")
@school_required
def backup_download(name):
    safe=FilePath(name).name
    return send_from_directory(str(BACKUP_DIR),safe,as_attachment=True)

@app.post("/backups/delete/<path:name>")
@school_required
def backup_delete(name):
    f=BACKUP_DIR/FilePath(name).name
    if f.exists() and f.is_file(): f.unlink(); flash("Backup deleted.","success")
    return redirect(url_for("automatic_backups_page"))

@app.route("/whatsapp-settings",methods=["GET","POST"])
@school_required
def whatsapp_settings():
    iid=session["institute_id"]
    inst=Institute.query.get(iid)
    cfg=WhatsAppConfig.query.filter_by(institute_id=iid).first()
    if not cfg:
        cfg=WhatsAppConfig(institute_id=iid,reminder_day=5)
        db.session.add(cfg); db.session.commit()
    if request.method=="POST":
        cfg.reminder_day=5
        cfg.enabled=request.form.get("enabled")=="1"
        # Advanced Meta connection values are saved per institute.
        if request.form.get("phone_number_id") is not None:
            cfg.phone_number_id=(request.form.get("phone_number_id") or "").strip()
        token=(request.form.get("access_token") or "").strip()
        if token: cfg.access_token=protect_secret(token)
        if request.form.get("api_version"):
            cfg.api_version=request.form["api_version"].strip()
        if request.form.get("template_name"):
            cfg.template_name=request.form["template_name"].strip()
        if request.form.get("template_language"):
            cfg.template_language=request.form["template_language"].strip()
        db.session.commit()
        flash("WhatsApp fee reminder settings saved for this institute.","success")
        return redirect(url_for("whatsapp_settings"))
    logs=FeeReminderLog.query.filter_by(institute_id=iid).order_by(FeeReminderLog.id.desc()).limit(100).all()
    connected=bool(cfg.phone_number_id and cfg.access_token)
    return render_template("whatsapp_settings.html",cfg=cfg,inst=inst,
        brand_phone=institute_brand_phone(inst),connected=connected,logs=logs)

@app.post("/whatsapp-reminders/test-run")
@school_required
def whatsapp_test_run():
    sent,failed,skipped=run_monthly_fee_reminders(force=True,institute_id=session["institute_id"])
    flash(f"Test completed: {sent} sent, {failed} failed, {skipped} skipped.","success" if not failed else "warning")
    return redirect(url_for("whatsapp_settings"))

@app.route("/fee-reminders")
@school_required
def fee_reminders():
    iid=session["institute_id"]
    inst=Institute.query.get(iid)
    cls=(request.args.get("class_name") or "").strip()
    month=(request.args.get("month") or "").strip()
    q=(request.args.get("q") or "").strip()

    fq=Fee.query.filter_by(institute_id=iid)
    if cls:
        fq=fq.join(Student,Fee.student_id==Student.id).filter(Student.class_name==cls)
    if month:
        fq=fq.filter(Fee.month==month)
    if q:
        if not cls:
            fq=fq.join(Student,Fee.student_id==Student.id)
        fq=fq.filter(Student.name.ilike(f"%{q}%"))

    rows=[]
    for f in fq.order_by(Fee.id.desc()).all():
        due=max(0,float(f.amount or 0)-float(f.discount or 0)-float(f.paid or 0))
        if due<=0:
            continue
        st=f.student
        phone="".join(ch for ch in (st.phone or "") if ch.isdigit())
        if phone.startswith("0") and len(phone)>=10:
            phone="92"+phone[1:]
        school=inst.name if inst else "School"
        msg=(f"Dear Parent/Guardian, fee reminder for {st.name} "
             f"(Admission No: {st.admission_no}, Class: {st.class_name}). "
             f"Fee Month: {f.month}. Outstanding amount: Rs. {due:,.0f}. "
             f"Kindly submit the pending fee. Thank you. — {school}")
        from urllib.parse import quote
        wa_url=f"https://wa.me/{phone}?text={quote(msg)}" if phone else ""
        rows.append({"fee":f,"student":st,"due":due,"phone":phone,
                     "message":msg,"wa_url":wa_url})

    classes=[x[0] for x in db.session.query(Student.class_name).filter_by(
        institute_id=iid,status="Active").distinct().order_by(Student.class_name).all() if x[0]]
    months=[x[0] for x in db.session.query(Fee.month).filter_by(
        institute_id=iid).distinct().order_by(Fee.id.desc()).all() if x[0]]
    total_due=sum(x["due"] for x in rows)
    return render_template("fee_reminders.html",inst=inst,rows=rows,classes=classes,
        months=months,selected_class=cls,selected_month=month,q=q,total_due=total_due)

@app.route("/accounts",methods=["GET","POST"])
@school_required
def accounts():
    iid=session["institute_id"]
    if request.method=="POST":
        db.session.add(Transaction(institute_id=iid,kind=request.form["kind"],category=request.form.get("category",""),description=request.form.get("description",""),
          amount=float(request.form.get("amount") or 0),txn_date=datetime.strptime(request.form["txn_date"],"%Y-%m-%d").date()))
        db.session.commit(); return redirect(url_for("accounts"))
    return render_template("accounts.html",tx=Transaction.query.filter_by(institute_id=iid).order_by(Transaction.txn_date.desc(),Transaction.id.desc()).all())


@app.route("/salaries",methods=["GET","POST"])
@school_required
def salaries():
    iid=session["institute_id"]
    if request.method=="POST":
        paid=float(request.form.get("paid") or 0)
        db.session.add(StaffSalary(institute_id=iid,staff_name=request.form["staff_name"].strip(),
            designation=request.form.get("designation",""),month=request.form["month"],
            salary=float(request.form.get("salary") or 0),paid=paid,
            paid_on=date.today() if paid>0 else None,note=request.form.get("note","")))
        db.session.commit(); flash("Staff salary entry saved.","success"); return redirect(url_for("salaries"))
    month=request.args.get("month","").strip()
    q=StaffSalary.query.filter_by(institute_id=iid)
    if month: q=q.filter_by(month=month)
    rows=q.order_by(StaffSalary.id.desc()).all()
    months=[r[0] for r in db.session.query(StaffSalary.month).filter_by(institute_id=iid).distinct().order_by(StaffSalary.id.desc()).all() if r[0]]
    return render_template("salaries.html",rows=rows,months=months,selected_month=month)

@app.post("/salary/<int:sid>/delete")
@school_required
def delete_salary(sid):
    r=StaffSalary.query.filter_by(id=sid,institute_id=session["institute_id"]).first_or_404()
    db.session.delete(r); db.session.commit(); return redirect(url_for("salaries"))

def branded_pdf_header(story, institute, title, subtitle=""):
    st=getSampleStyleSheet()
    if institute.logo:
        from reportlab.platypus import Image
        logo_path=UPLOAD_DIR/institute.logo
        if logo_path.exists():
            im=Image(str(logo_path),width=50,height=50); im.hAlign="CENTER"; story.append(im)
    story.append(Paragraph(institute.name,st["Title"]))
    story.append(Paragraph(title,st["Heading2"]))
    if subtitle: story.append(Paragraph(subtitle,st["Normal"]))
    story.append(Spacer(1,10))
    return st

@app.route("/accounts/expenses.pdf")
@school_required
def expenses_pdf():
    iid=session["institute_id"]; i=current_inst()
    rows=Transaction.query.filter_by(institute_id=iid,kind="Expense").order_by(Transaction.txn_date,Transaction.id).all()
    total=sum(r.amount for r in rows)
    out=io.BytesIO(); doc=SimpleDocTemplate(out,pagesize=A4,rightMargin=22,leftMargin=22,topMargin=24,bottomMargin=24)
    story=[]; st=branded_pdf_header(story,i,"Expense Statement",f"Generated: {date.today().isoformat()}")
    data=[["Date","Category","Description","Amount"]]+[[str(r.txn_date),r.category,r.description,f"Rs {r.amount:,.2f}"] for r in rows]
    data.append(["","","TOTAL EXPENSE",f"Rs {total:,.2f}"])
    t=Table(data,repeatRows=1,colWidths=[70,95,245,95])
    t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor(i.primary_color)),("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
        ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#eef3f8")),("GRID",(0,0),(-1,-1),.35,colors.grey),
        ("FONTSIZE",(0,0),(-1,-1),8),("ALIGN",(-1,1),(-1,-1),"RIGHT"),("VALIGN",(0,0),(-1,-1),"MIDDLE")]))
    story += [t,Spacer(1,12),Paragraph(f"<b>Total Expense: Rs {total:,.2f}</b>",st["Heading3"]),
              Paragraph("Powered by E-2 Solutions | Contact: +923010012627",e2_pdf_footer_style(st))]
    doc.build(story); out.seek(0)
    return send_file(out,as_attachment=True,download_name="expense_statement.pdf",mimetype="application/pdf")

@app.route("/accounts/summary.pdf")
@school_required
def accounts_pdf():
    iid=session["institute_id"]; i=current_inst()
    rows=Transaction.query.filter_by(institute_id=iid).order_by(Transaction.txn_date,Transaction.id).all()
    fee_paid=sum(f.paid for f in Fee.query.filter_by(institute_id=iid).all())
    income=sum(r.amount for r in rows if r.kind=="Income")
    expense=sum(r.amount for r in rows if r.kind=="Expense")
    salary_paid=sum(r.paid for r in StaffSalary.query.filter_by(institute_id=iid).all())
    net=fee_paid+income-expense-salary_paid
    out=io.BytesIO(); doc=SimpleDocTemplate(out,pagesize=A4,rightMargin=22,leftMargin=22,topMargin=24,bottomMargin=24)
    story=[]; st=branded_pdf_header(story,i,"Accounts Summary",f"Generated: {date.today().isoformat()}")
    summary=[["Particular","Total"],["Fee Received",f"Rs {fee_paid:,.2f}"],["Other Income",f"Rs {income:,.2f}"],
             ["Expenses",f"Rs {expense:,.2f}"],["Staff Salaries Paid",f"Rs {salary_paid:,.2f}"],["Net Balance",f"Rs {net:,.2f}"]]
    t=Table(summary,colWidths=[280,180])
    t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor(i.primary_color)),("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),("GRID",(0,0),(-1,-1),.5,colors.grey),
        ("FONTSIZE",(0,0),(-1,-1),10),("ALIGN",(1,1),(1,-1),"RIGHT"),("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#eef3f8"))]))
    story += [t,Spacer(1,16),Paragraph("Powered by E-2 Solutions | Contact: +923010012627",e2_pdf_footer_style(st))]
    doc.build(story); out.seek(0)
    return send_file(out,as_attachment=True,download_name="accounts_summary.pdf",mimetype="application/pdf")

@app.route("/salaries.pdf")
@school_required
def salaries_pdf():
    iid=session["institute_id"]; i=current_inst(); month=request.args.get("month","").strip()
    q=StaffSalary.query.filter_by(institute_id=iid)
    if month: q=q.filter_by(month=month)
    rows=q.order_by(StaffSalary.staff_name).all()
    total_salary=sum(r.salary for r in rows); total_paid=sum(r.paid for r in rows); total_due=sum(max(0,r.salary-r.paid) for r in rows)
    out=io.BytesIO(); doc=SimpleDocTemplate(out,pagesize=A4,rightMargin=18,leftMargin=18,topMargin=24,bottomMargin=24)
    story=[]; st=branded_pdf_header(story,i,"Staff Salary Statement",f"Month: {month or 'All'}")
    data=[["Staff","Designation","Month","Salary","Paid","Due","Status"]]
    for r in rows:
        due=max(0,r.salary-r.paid); status="Paid" if due<=0 and r.paid>0 else ("Partial" if r.paid>0 else "Unpaid")
        data.append([r.staff_name,r.designation,r.month,f"{r.salary:,.0f}",f"{r.paid:,.0f}",f"{due:,.0f}",status])
    data.append(["","","TOTAL",f"{total_salary:,.0f}",f"{total_paid:,.0f}",f"{total_due:,.0f}",""])
    t=Table(data,repeatRows=1,colWidths=[90,80,70,65,65,65,55])
    t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor(i.primary_color)),("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
        ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#eef3f8")),("GRID",(0,0),(-1,-1),.35,colors.grey),
        ("FONTSIZE",(0,0),(-1,-1),7.5),("ALIGN",(3,1),(5,-1),"RIGHT")]))
    story += [t,Spacer(1,12),Paragraph(f"<b>Total Salary: Rs {total_salary:,.2f} | Paid: Rs {total_paid:,.2f} | Due: Rs {total_due:,.2f}</b>",st["Heading3"]),
              Paragraph("Powered by E-2 Solutions | Contact: +923010012627",e2_pdf_footer_style(st))]
    doc.build(story); out.seek(0)
    return send_file(out,as_attachment=True,download_name=f"staff_salaries_{month or 'all'}.pdf",mimetype="application/pdf")

@app.route("/exams",methods=["GET","POST"])
@school_required
def exams():
    iid=session["institute_id"]
    classes=sorted({(" ".join(x[0].split())) for x in db.session.query(Student.class_name).filter_by(institute_id=iid,status="Active").all() if x[0]},key=str.casefold)
    cls=request.args.get("class_name","").strip()
    if request.method=="POST":
        action=request.form.get("action","exam")
        if action=="subject":
            cname=request.form["class_name"].strip(); sub=request.form["subject"].strip()
            exists=ClassSubject.query.filter_by(institute_id=iid,class_name=cname,subject=sub).first()
            if not exists:
                db.session.add(ClassSubject(institute_id=iid,class_name=cname,subject=sub,total_marks=float(request.form.get("total_marks") or 100)))
                db.session.commit()
            flash("Class subject saved.","success")
            return redirect(url_for("exams",class_name=cname))
        cname=request.form["class_name"].strip()
        if ClassSubject.query.filter_by(institute_id=iid,class_name=cname).count()==0:
            flash("First add subjects for this class.","error"); return redirect(url_for("exams",class_name=cname))
        db.session.add(ExamSession(institute_id=iid,name=request.form["name"].strip(),class_name=cname,
                                   exam_date=datetime.strptime(request.form["exam_date"],"%Y-%m-%d").date()))
        db.session.commit(); flash("Exam created once for the whole class.","success")
        return redirect(url_for("exams",class_name=cname))
    subjects=ClassSubject.query.filter_by(institute_id=iid,class_name=cls).order_by(ClassSubject.subject).all() if cls else []
    q=ExamSession.query.filter_by(institute_id=iid)
    if cls: q=q.filter_by(class_name=cls)
    exam_sessions=q.order_by(ExamSession.exam_date.desc(),ExamSession.id.desc()).all()
    return render_template("exams.html",classes=classes,subjects=subjects,exam_sessions=exam_sessions,selected_class=cls)

@app.post("/class-subject/<int:sid>/delete")
@school_required
def delete_class_subject(sid):
    x=ClassSubject.query.filter_by(id=sid,institute_id=session["institute_id"]).first_or_404()
    db.session.delete(x); db.session.commit()
    return redirect(url_for("exams",class_name=x.class_name))

@app.route("/exam-session/<int:eid>")
@school_required
def exam_session_students(eid):
    iid=session["institute_id"]; e=ExamSession.query.filter_by(id=eid,institute_id=iid).first_or_404()
    students=Student.query.filter_by(institute_id=iid,class_name=e.class_name,status="Active").order_by(Student.name).all()
    subjects=ClassSubject.query.filter_by(institute_id=iid,class_name=e.class_name).order_by(ClassSubject.subject).all()
    return render_template("exam_students.html",exam=e,students=students,subjects=subjects)

@app.route("/exam-session/<int:eid>/student/<int:sid>",methods=["GET","POST"])
@school_required
def student_exam_marks(eid,sid):
    iid=session["institute_id"]
    e=ExamSession.query.filter_by(id=eid,institute_id=iid).first_or_404()
    s=Student.query.filter_by(id=sid,institute_id=iid,class_name=e.class_name).first_or_404()
    subjects=ClassSubject.query.filter_by(institute_id=iid,class_name=e.class_name).order_by(ClassSubject.subject).all()
    if request.method=="POST":
        for sub in subjects:
            raw=request.form.get(f"marks_{sub.id}","").strip()
            if raw=="":
                continue
            om=max(0,min(float(raw),sub.total_marks))
            m=ExamMark.query.filter_by(institute_id=iid,exam_session_id=e.id,student_id=s.id,class_subject_id=sub.id).first()
            if not m:
                m=ExamMark(institute_id=iid,exam_session_id=e.id,student_id=s.id,class_subject_id=sub.id)
                db.session.add(m)
            m.obtained_marks=om
        db.session.commit(); flash("Student marks saved.","success")
        return redirect(url_for("exam_session_students",eid=e.id))
    existing={m.class_subject_id:m for m in ExamMark.query.filter_by(institute_id=iid,exam_session_id=e.id,student_id=s.id).all()}
    return render_template("student_exam_marks.html",exam=e,student=s,subjects=subjects,existing=existing)

def pro_result_data(iid,eid,sid=None):
    e=ExamSession.query.filter_by(id=eid,institute_id=iid).first_or_404()
    subjects=ClassSubject.query.filter_by(institute_id=iid,class_name=e.class_name).order_by(ClassSubject.subject).all()
    students=[Student.query.filter_by(id=sid,institute_id=iid,class_name=e.class_name).first_or_404()] if sid else Student.query.filter_by(institute_id=iid,class_name=e.class_name,status="Active").order_by(Student.name).all()
    rows=[]
    for s in students:
        marks={m.class_subject_id:m.obtained_marks for m in ExamMark.query.filter_by(institute_id=iid,exam_session_id=e.id,student_id=s.id).all()}
        detail=[]; total=0; obtained=0
        for sub in subjects:
            om=marks.get(sub.id,0); total+=sub.total_marks; obtained+=om
            detail.append((sub.subject,sub.total_marks,om))
        pct=(obtained/total*100) if total else 0
        grade="A+" if pct>=90 else "A" if pct>=80 else "B+" if pct>=70 else "B" if pct>=60 else "C" if pct>=50 else "D" if pct>=40 else "F"
        rows.append((s,detail,total,obtained,pct,grade))
    return e,subjects,rows

def result_watermark(canvas,doc,institute):
    canvas.saveState()
    if institute.logo:
        lp=UPLOAD_DIR/institute.logo
        if lp.exists():
            try:
                canvas.setFillAlpha(.06); canvas.setStrokeAlpha(.06)
                size=260; x=(A4[0]-size)/2; y=(A4[1]-size)/2-20
                canvas.drawImage(ImageReader(str(lp)),x,y,width=size,height=size,mask="auto",preserveAspectRatio=True,anchor="c")
            except Exception:
                pass
    canvas.restoreState()

@app.route("/exam-session/<int:eid>/student/<int:sid>.pdf")
@school_required
def professional_student_result_pdf(eid,sid):
    iid=session["institute_id"]; i=current_inst(); e,subjects,rows=pro_result_data(iid,eid,sid)
    s,detail,total,obtained,pct,grade=rows[0]
    out=io.BytesIO(); doc=SimpleDocTemplate(out,pagesize=A4,rightMargin=28,leftMargin=28,topMargin=22,bottomMargin=28)
    st=getSampleStyleSheet(); story=[]
    brand=colors.HexColor(i.primary_color)
    if i.logo:
        from reportlab.platypus import Image
        lp=UPLOAD_DIR/i.logo
        if lp.exists():
            im=Image(str(lp),width=62,height=62); im.hAlign="CENTER"; story.append(im)
    title_style=st["Title"]; title_style.fontSize=19; title_style.leading=22
    story += [Paragraph(i.name,title_style),Paragraph("<b>ACADEMIC RESULT REPORT</b>",st["Heading2"]),
              Paragraph(f"{e.name} &nbsp; | &nbsp; Class: {e.class_name} &nbsp; | &nbsp; Date: {e.exam_date.strftime('%d-%m-%Y')}",st["Normal"]),Spacer(1,10)]
    info=[[Paragraph(f"<b>Student Name:</b> {s.name}",st["Normal"]),Paragraph(f"<b>Admission No:</b> {s.admission_no}",st["Normal"])],
          [Paragraph(f"<b>Guardian:</b> {s.father_name}",st["Normal"]),Paragraph(f"<b>Section:</b> {s.section}",st["Normal"])]]
    it=Table(info,colWidths=[265,265]); it.setStyle(TableStyle([("BOX",(0,0),(-1,-1),.5,colors.grey),("INNERGRID",(0,0),(-1,-1),.25,colors.lightgrey),("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#f7f9fc")),("PADDING",(0,0),(-1,-1),6)]))
    story += [it,Spacer(1,12)]
    data=[["Subject","Total Marks","Obtained Marks","%","Grade"]]
    for sub,tm,om in detail:
        sp=(om/tm*100) if tm else 0
        sg="A+" if sp>=90 else "A" if sp>=80 else "B+" if sp>=70 else "B" if sp>=60 else "C" if sp>=50 else "D" if sp>=40 else "F"
        data.append([sub,f"{tm:.0f}",f"{om:.0f}",f"{sp:.1f}%",sg])
    data.append(["GRAND TOTAL",f"{total:.0f}",f"{obtained:.0f}",f"{pct:.2f}%",grade])
    mt=Table(data,repeatRows=1,colWidths=[225,85,100,65,55])
    mt.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),brand),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#eaf0f8")),("GRID",(0,0),(-1,-1),.5,colors.grey),
        ("ALIGN",(1,1),(-1,-1),"CENTER"),("FONTSIZE",(0,0),(-1,-1),8.5),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)]))
    story += [mt,Spacer(1,10)]
    summary=Table([["Total Marks",f"{total:.0f}","Obtained",f"{obtained:.0f}","Percentage",f"{pct:.2f}%","Grade",grade]],colWidths=[72,55,65,55,72,62,45,45])
    summary.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#f2f5fa")),("GRID",(0,0),(-1,-1),.45,colors.grey),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("ALIGN",(0,0),(-1,-1),"CENTER"),("FONTSIZE",(0,0),(-1,-1),8)]))
    story += [summary,Spacer(1,16),
              Table([["Class Teacher Signature","","Principal Signature"],["____________________","","____________________"]],colWidths=[180,170,180],style=[("ALIGN",(0,0),(-1,-1),"CENTER"),("FONTSIZE",(0,0),(-1,-1),8)]),
              Spacer(1,12),
              Paragraph("<b>Grading Scale:</b> A+ 90-100 | A 80-89 | B+ 70-79 | B 60-69 | C 50-59 | D 40-49 | F Below 40",st["Normal"]),
              Paragraph("Powered by E-2 Solutions | Contact: +923010012627",e2_pdf_footer_style(st))]
    doc.build(story,onFirstPage=lambda c,d:result_watermark(c,d,i),onLaterPages=lambda c,d:result_watermark(c,d,i)); out.seek(0)
    return send_file(out,as_attachment=True,download_name=f"{s.name}_{e.name}_result.pdf",mimetype="application/pdf")

@app.route("/exam-session/<int:eid>/grand-sheet.pdf")
@school_required
def professional_grand_sheet_pdf(eid):
    iid=session["institute_id"]; i=current_inst(); e,subjects,rows=pro_result_data(iid,eid)
    out=io.BytesIO(); doc=SimpleDocTemplate(out,pagesize=landscape(A4),rightMargin=16,leftMargin=16,topMargin=20,bottomMargin=22)
    st=getSampleStyleSheet(); story=[Paragraph(i.name,st["Title"]),Paragraph(f"{e.name} - Class {e.class_name} Grand Result Sheet",st["Heading2"]),Spacer(1,8)]
    data=[["Adm #","Student"]+[s.subject for s in subjects]+["Total","Obt.","%","Grade"]]
    for student,detail,total,obtained,pct,grade in rows:
        data.append([student.admission_no,student.name]+[f"{om:.0f}/{tm:.0f}" for _,tm,om in detail]+[f"{total:.0f}",f"{obtained:.0f}",f"{pct:.1f}%",grade])
    pw=landscape(A4)[0]-32; fixed=58+105+48+52+48+42; sw=max(40,(pw-fixed)/max(1,len(subjects)))
    tb=Table(data,repeatRows=1,colWidths=[58,105]+[sw]*len(subjects)+[48,52,48,42])
    tb.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor(i.primary_color)),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("GRID",(0,0),(-1,-1),.35,colors.grey),("FONTSIZE",(0,0),(-1,-1),6.8),("ALIGN",(2,1),(-1,-1),"CENTER"),("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f6f8fb")])]))
    avg=sum(r[4] for r in rows)/len(rows) if rows else 0
    story += [tb,Spacer(1,8),Paragraph(f"<b>Total Students: {len(rows)} | Class Average: {avg:.2f}%</b>",st["Normal"]),Paragraph("Powered by E-2 Solutions | Contact: +923010012627",e2_pdf_footer_style(st))]
    doc.build(story); out.seek(0)
    return send_file(out,as_attachment=True,download_name=f"{e.class_name}_{e.name}_grand_sheet.pdf",mimetype="application/pdf")


def migrate_erp22_content_bank():
    """Upgrade older Lecture tables using portable SQLAlchemy inspection."""
    try:
        cols=_table_columns("lecture")
        if not cols:
            return
        additions={
            "chapter":"VARCHAR(160) DEFAULT ''",
            "content_type":"VARCHAR(30) DEFAULT 'Video'",
            "video_url":"VARCHAR(700) DEFAULT ''",
            "file_name":"VARCHAR(255) DEFAULT ''",
            "description":"TEXT DEFAULT ''",
            "created_on":"DATE"
        }
        changed=False
        for name,kind in additions.items():
            if name not in cols:
                db.session.execute(db.text(f'ALTER TABLE lecture ADD COLUMN "{name}" {kind}'))
                changed=True
        if changed:
            db.session.commit()
        print("[ERP22.2] Content Bank database schema: OK")
    except Exception as exc:
        db.session.rollback()
        print("[ERP22.2] Content Bank migration skipped/failed:",repr(exc))

def seed():
    db.create_all()
    if not User.query.filter_by(email="admin@e2solutions.local").first():
        db.session.add(User(email="admin@e2solutions.local",password_hash=generate_password_hash("admin123"),role="superadmin"))
        db.session.commit()

    # First-run demo environment: if there is no school at all, create a clearly
    # labelled demo school so the dashboard is never an empty shell. Existing
    # institutes/data are never replaced or deleted.
    if Institute.query.count() == 0:
        inst=Institute(
            name="E-2 Demo School", email="demo@e2solutions.local",
            tagline="Professional School Management", phone="+92 300 0000000",
            address="Demo Campus", primary_color="#2457d6", active=True)
        db.session.add(inst); db.session.flush()
        school_user=User(
            institute_id=inst.id, email="demo@e2solutions.local",
            password_hash=generate_password_hash("school123"), role="school_admin")
        db.session.add(school_user); db.session.flush()

        demo_students=[
            ("E2-001","Ayesha Khan","Muhammad Khan","0300-1111111","Play Group","A",8500),
            ("E2-002","Hamza Ahmed","Tariq Ahmed","0300-2222222","Class 1","A",9000),
            ("E2-003","Maham Ali","Sajid Ali","0300-3333333","Class 2","B",9500),
            ("E2-004","Hassan Raza","Raza Hussain","0300-4444444","Class 3","A",10000),
            ("E2-005","Zoya Malik","Imran Malik","0300-5555555","Class 4","B",10500),
            ("E2-006","Abdullah Noor","Noor Ahmed","0300-6666666","Class 5","A",11000),
        ]
        students=[]
        for adm,name,father,phone,cls,sec,fee in demo_students:
            st=Student(institute_id=inst.id,admission_no=adm,name=name,father_name=father,
                       phone=phone,class_name=cls,section=sec,monthly_fee=fee,status="Active")
            db.session.add(st); db.session.flush(); students.append(st)
        today=date.today()
        for idx,st in enumerate(students):
            amount=st.monthly_fee
            paid=amount if idx % 3 != 0 else amount*0.55
            db.session.add(Fee(institute_id=inst.id,student_id=st.id,month=today.strftime("%B %Y"),
                               amount=amount,discount=0,paid=paid,paid_on=today if paid else None))
            db.session.add(Attendance(institute_id=inst.id,student_id=st.id,attendance_date=today,
                                      status="Absent" if idx==4 else "Present",marked_by=school_user.id))
        db.session.add(Transaction(institute_id=inst.id,kind="Income",category="Admission Fees",description="Demo fee collection",amount=72000,txn_date=today))
        db.session.add(Transaction(institute_id=inst.id,kind="Expense",category="Utilities",description="Demo monthly utilities",amount=18500,txn_date=today))
        teacher_user=User(institute_id=inst.id,email="teacher@e2solutions.local",password_hash=generate_password_hash("teacher123"),role="teacher")
        db.session.add(teacher_user); db.session.flush()
        db.session.add(TeacherProfile(institute_id=inst.id,user_id=teacher_user.id,employee_no="T-001",full_name="Demo Teacher",
                                      email="teacher@e2solutions.local",designation="Senior Teacher",status="Active",monthly_salary=45000,joining_date=today))
        exam=ExamSession(institute_id=inst.id,name="First Term Assessment",class_name="Class 5",exam_date=today)
        db.session.add(exam); db.session.flush()
        for subj in ["English","Mathematics","Science","Computer"]:
            db.session.add(ClassSubject(institute_id=inst.id,class_name="Class 5",subject=subj,total_marks=100))
        db.session.add(Lecture(institute_id=inst.id,class_name="Class 5",subject="Computer",chapter="Introduction",
                               title="Welcome to Computer Studies",description="Demo learning resource",content_type="Video",created_on=today))
        db.session.commit()
        print("[E2] First-run demo school created: demo@e2solutions.local / school123")

# Recover an older project-local DB before SQLAlchemy first creates a fresh one.
recover_legacy_database(DB_PATH)
with app.app_context():
    seed()
    migrate_db()
    migrate_institute_branding()
    migrate_erp22_content_bank()
    db.create_all()

def open_browser():
    if os.environ.get("E2_NO_BROWSER") == "1":
        return
    webbrowser.open(f"http://127.0.0.1:{os.environ.get('E2_PORT','5190')}/")



@app.route("/class-view/<path:class_name>")
@login_required
def class_view_safe(class_name):
    iid=session.get("institute_id")
    if not iid: return redirect(url_for("login"))
    rows=Student.query.filter_by(institute_id=iid,class_name=class_name).order_by(Student.name).all()
    inst=Institute.query.get(iid)
    return render_template("class_view_safe.html",rows=rows,class_name=class_name,institute=inst)

@app.route("/analytics")
@school_required
def analytics_dashboard():
    iid=session["institute_id"]
    students=Student.query.filter_by(institute_id=iid).all()
    active=[x for x in students if (x.status or "").strip().lower()=="active"]
    classes={}
    for x in active:
        c=" ".join((x.class_name or "Unassigned").split())
        classes[c]=classes.get(c,0)+1

    attendance=Attendance.query.filter_by(institute_id=iid).all()
    att_total=len(attendance)
    present=sum(1 for x in attendance if (x.status or "").strip().lower()=="present")
    attendance_pct=round(present*100/att_total,1) if att_total else 0

    fees=Fee.query.filter_by(institute_id=iid).all()
    fee_received=sum(float(x.paid or 0) for x in fees)
    fee_pending=sum(max(0,float(x.amount or 0)-float(x.discount or 0)-float(x.paid or 0)) for x in fees)

    tx=Transaction.query.filter_by(institute_id=iid).all()
    income=sum(float(x.amount or 0) for x in tx if (x.kind or "").lower()=="income")
    expenses=sum(float(x.amount or 0) for x in tx if (x.kind or "").lower()=="expense")

    inst=Institute.query.get(iid)
    return render_template("analytics.html",inst=inst,institute=inst,
        total_students=len(students),active_students=len(active),classes=classes,
        attendance_pct=attendance_pct,attendance_total=att_total,
        fee_received=fee_received,fee_pending=fee_pending,
        income=income,expenses=expenses,
        result_sessions=ExamSession.query.filter_by(institute_id=iid).count(),
        lectures=Lecture.query.filter_by(institute_id=iid).count())

# Mobile/API layer is registered after all models and migrations are ready.
# It is intentionally additive: the existing web/desktop ERP routes remain unchanged.
try:
    from mobile_api import register_mobile_api
    register_mobile_api(app, db, {
        "Institute": Institute, "User": User, "Student": Student, "Fee": Fee,
        "Attendance": Attendance, "Transaction": Transaction, "ExamSession": ExamSession,
    })
    print("[E2] Mobile API: READY")
except Exception as _api_exc:
    print("[E2] Mobile API registration warning:", _api_exc)



@app.route("/mobile-app")
@login_required
def mobile_app():
    iid=session.get("institute_id")
    inst=Institute.query.get(iid) if iid else None
    students=Student.query.filter_by(institute_id=iid,status="Active").count() if iid else 0
    teachers=TeacherProfile.query.filter_by(institute_id=iid,status="Active").count() if iid else 0
    fees=Fee.query.filter_by(institute_id=iid).all() if iid else []
    tx=Transaction.query.filter_by(institute_id=iid).all() if iid else []
    attendance=Attendance.query.filter_by(institute_id=iid,attendance_date=date.today()).all() if iid else []
    present=sum(1 for a in attendance if (a.status or "").strip().lower()=="present")
    attendance_pct=round(present*100/len(attendance),1) if attendance else 0
    collected=sum(float(f.paid or 0) for f in fees)
    due=sum(max(0,float(f.amount or 0)-float(f.discount or 0)-float(f.paid or 0)) for f in fees)
    income=sum(float(x.amount or 0) for x in tx if (x.kind or "").lower()=="income")
    expense=sum(float(x.amount or 0) for x in tx if (x.kind or "").lower()=="expense")
    profit=income-expense
    exams=ExamSession.query.filter_by(institute_id=iid).count() if iid else 0
    user=User.query.get(session.get("uid")) if session.get("uid") else None
    user_name=(user.email.split('@')[0].replace('.',' ').replace('_',' ').title() if user else 'Administrator')
    return render_template("mobile_app.html",inst=inst,total_students=students,teacher_count=teachers,
        collected=collected,due=due,income=income,expense=expense,profit=profit,present_today=present,
        attendance_today_count=len(attendance),attendance_today_pct=attendance_pct,exam_count=exams,user_name=user_name)

@app.route("/mobile.webmanifest")
def mobile_manifest():
    return jsonify({
        "name":"E-2 School ERP","short_name":"E-2 ERP","start_url":"/mobile-app","scope":"/",
        "display":"standalone","background_color":"#f4f7fb","theme_color":"#123a8c",
        "description":"E-2 School ERP mobile application",
        "icons":[
            {"src":"/static/icons/icon-192.png","sizes":"192x192","type":"image/png","purpose":"any maskable"},
            {"src":"/static/icons/icon-512.png","sizes":"512x512","type":"image/png","purpose":"any maskable"}
        ]
    })

if __name__=="__main__":
    import os, socket
    host = os.environ.get("E2_HOST", "0.0.0.0")
    port = int(os.environ.get("E2_PORT", "5190"))
    local_ip = "127.0.0.1"
    try:
        # UDP connect does not send traffic; it lets Windows select the
        # active LAN interface, which is more reliable than gethostbyname().
        sock=socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.connect(("8.8.8.8",80))
        local_ip=sock.getsockname()[0]
        sock.close()
    except Exception:
        try:
            local_ip = socket.gethostbyname(socket.gethostname())
        except Exception:
            pass
    threading.Timer(1.2,open_browser).start()
    print(f"E-2 School ERP STARTING ON http://127.0.0.1:{port}/")
    print(f"MOBILE TEST URL: http://{local_ip}:{port}/")
    print("For phone testing, keep PC and phone on the same network.")
    app.run(debug=False,host=host,port=port,use_reloader=False)
